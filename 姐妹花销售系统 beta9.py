import base64
import configparser
import ctypes
import glob
import json
import logging
import os
import platform
import re
import shutil
import sqlite3
import sys
import threading
import time
import tkinter as tk
import warnings
from ctypes import windll
from ctypes import wintypes
from datetime import datetime, date, timedelta
from tkinter import filedialog
from tkinter import messagebox
from tkinter import ttk, font, scrolledtext

import matplotlib.pyplot as plt
import pygame
import pyttsx3
import requests
import test_server
import win32api
import win32con
import win32event
import win32gui
from PIL import Image, ImageDraw, ImageTk
from cryptography.hazmat.primitives.ciphers.aead import AESGCM
from lunar_python import Solar
from openpyxl import Workbook
from pystray import Icon, Menu, MenuItem

from config import THEME_CONFIG, DB_PATH, SCALE_FACTOR

# ---------------- 工具函数 ----------------
user32 = ctypes.WinDLL('user32', use_last_error=True)
user32.SetTimer.argtypes = [wintypes.HWND, wintypes.UINT, wintypes.UINT, wintypes.LPVOID]
user32.SetTimer.restype = wintypes.UINT
user32.KillTimer.argtypes = [wintypes.HWND, wintypes.UINT]
user32.KillTimer.restype = wintypes.BOOL

def button_animation(button):
    """按钮点击动画：短暂变色+轻微抖动"""
    original_style = button.cget("style")
    original_pos = button.place_info() or button.grid_info()  # 兼容不同布局

    # 临时变色
    button.config(style="success.TButton")

    # 轻微抖动效果
    def shake(offset=2):
        if button.winfo_exists():  # 检查按钮是否仍存在
            if 'row' in original_pos:  # grid布局
                button.grid(row=original_pos['row'],
                            column=original_pos['column'],
                            padx=original_pos.get('padx', 0) + offset)
            elif 'x' in original_pos:  # place布局
                button.place(x=int(original_pos['x']) + offset,
                             y=original_pos['y'])
            button.after(50, lambda: shake(-offset) if offset != 0 else reset)

    def reset():
        if button.winfo_exists():
            button.config(style=original_style)
            if 'row' in original_pos:  # 恢复grid布局
                button.grid(row=original_pos['row'],
                            column=original_pos['column'],
                            padx=original_pos.get('padx', 0))
            elif 'x' in original_pos:  # 恢复place布局
                button.place(x=original_pos['x'], y=original_pos['y'])
            # 显示确认提示
            show_temp_message(button, "已确认！")

    shake()

def show_temp_message(widget, text, duration=1000):
    """在控件旁显示临时消息"""
    msg_window = tk.Toplevel(widget)
    msg_window.overrideredirect(True)  # 无边框
    msg_window.attributes("-topmost", True)

    # 计算显示位置（控件右下方）
    x = widget.winfo_rootx() + widget.winfo_width() + 10
    y = widget.winfo_rooty() + widget.winfo_height() // 2

    msg_window.geometry(f"+{x}+{y}")
    ttk.Label(msg_window, text=text, style="info.TLabel").pack(padx=5, pady=2)

    # 自动关闭
    def close_msg():
        if msg_window.winfo_exists():
            msg_window.destroy()

    msg_window.after(duration, close_msg)

def resource_path(relative_path):
    """获取资源文件的绝对路径（适配开发环境和打包后环境）"""
    if getattr(sys, 'frozen', False):
        # 打包后的EXE环境：使用EXE所在目录作为基准路径
        base_path = os.path.dirname(sys.executable)
    else:
        # 开发环境：使用当前脚本所在目录作为基准路径
        base_path = os.path.dirname(os.path.abspath(__file__))
    return os.path.join(base_path, relative_path)

"""加载配置文件"""
def load_config():
    """加载配置文件"""
    config = configparser.ConfigParser()
    config_path = resource_path("config.ini")  # 使用修正后的路径
    # 若配置文件不存在则在程序运行目录创建默认配置
    if not os.path.exists(config_path):
        config["push"] = {
            "interval": "5",
            "endpoint": "https://58e4706c.tw.cpolar.io/pushdata"
        }
        config["database"] = {
            "last_push_table": "push_status"
        }
        config["backup"] = {
            "enabled": "false",
            "time": "03:00",
            "interval_minutes": "30",
            "path": os.path.join(resource_path(""), "backups")  # 绝对路径
        }
        with open(config_path, "w", encoding="utf-8") as f:
            config.write(f)
    config.read(config_path, encoding="utf-8")
    return config

# 新增保存配置的函数
def save_config_to_file():
    """保存配置到文件"""
    try:
        config_path = resource_path("config.ini")
        with open(config_path, "w", encoding="utf-8") as f:
            CONFIG.write(f)
        print("配置已成功保存到文件")
    except Exception as e:
        print(f"保存配置文件失败: {str(e)}")

# 加载配置（全局使用）
CONFIG = load_config()

def load_theme():
    """加载并返回上次保存的主题"""
    try:
        with open(resource_path(THEME_CONFIG), 'r', encoding='utf-8') as f:
            return f.read().strip()
    except:
        return None

def save_theme(theme_name):
    """保存当前主题"""
    try:
        with open(resource_path(THEME_CONFIG), 'w', encoding='utf-8') as f:
            f.write(theme_name)
    except Exception as e:
        print("Theme save error:", e)

# 添加配置存储函数（放在工具函数区域）
def save_exit_preference(choice, remember):
    """保存退出偏好设置"""
    config_path = resource_path("exit_preference.ini")
    config = configparser.ConfigParser()
    config.read(config_path)
    if not config.has_section("Exit"):
        config.add_section("Exit")
    config.set("Exit", "choice", choice)  # "exit" 或 "tray"
    config.set("Exit", "remember", str(remember))
    with open(config_path, "w", encoding="utf-8") as f:
        config.write(f)

def load_exit_preference():
    """加载退出偏好设置"""
    config_path = resource_path("exit_preference.ini")
    if not os.path.exists(config_path):
        return None, False
    config = configparser.ConfigParser()
    config.read(config_path)
    if not config.has_section("Exit"):
        return None, False
    choice = config.get("Exit", "choice", fallback=None)
    remember = config.getboolean("Exit", "remember", fallback=False)
    return choice, remember

def reset_exit_preference():
    """重置退出偏好设置"""
    config_path = resource_path("exit_preference.ini")
    if os.path.exists(config_path):
        os.remove(config_path)

def init_db():
    """初始化或升级数据库结构"""
    conn = sqlite3.connect(DB_PATH, timeout=30, check_same_thread=False)
    cur = conn.cursor()

    # users 表
    cur.execute("""
        CREATE TABLE IF NOT EXISTS users (
            username TEXT PRIMARY KEY,
            password TEXT NOT NULL,
            avatar TEXT
        )
    """)

    # 检查是否已有 avatar 字段（兼容老库）
    cur.execute("PRAGMA table_info(users)")
    cols = {r[1] for r in cur.fetchall()}
    if 'avatar' not in cols:
        cur.execute("ALTER TABLE users ADD COLUMN avatar TEXT")
        cur.execute("UPDATE users SET avatar = 'profile_photo.png' WHERE avatar IS NULL")

    # 检查用户表是否为空，若空则插入默认用户
    cur.execute("SELECT COUNT(*) FROM users")
    if cur.fetchone()[0] == 0:
        default_avatar = "profile_photo.png"
        cur.execute(
            "INSERT INTO users (username, password, avatar) VALUES (?, ?, ?)",
            ("a", "123", default_avatar)
        )

    # members 表
    cur.execute("""
        CREATE TABLE IF NOT EXISTS members (
            member_id INTEGER PRIMARY KEY AUTOINCREMENT,
            phone TEXT UNIQUE,
            balance REAL DEFAULT 0,
            remark TEXT,
            join_date TEXT
        )
    """)
    # 检查是否已有join_date字段（兼容老库）
    cur.execute("PRAGMA table_info(members)")
    cols = {r[1] for r in cur.fetchall()}
    if 'join_date' not in cols:
        cur.execute("ALTER TABLE members ADD COLUMN join_date TEXT")
        # 为现有会员设置默认日期
        cur.execute("UPDATE members SET join_date = datetime('now', 'localtime') WHERE join_date IS NULL")

    # inventory 表
    cur.execute("""
        CREATE TABLE IF NOT EXISTS inventory (
            item_id INTEGER PRIMARY KEY AUTOINCREMENT,
            category TEXT,
            name TEXT UNIQUE,
            price REAL DEFAULT 0,
            member_price REAL DEFAULT 0,
            remark TEXT
        )
    """)

    # 移除库存 stock 字段（兼容升级）
    cur.execute("PRAGMA table_info(inventory)")
    cols = {r[1] for r in cur.fetchall()}
    if 'stock' in cols:
        # SQLite 不支持直接 DROP COLUMN，这里需要迁移表
        cur.execute("""
            CREATE TABLE IF NOT EXISTS inventory_new (
                item_id INTEGER PRIMARY KEY AUTOINCREMENT,
                category TEXT,
                name TEXT UNIQUE,
                price REAL DEFAULT 0,
                member_price REAL DEFAULT 0,
                remark TEXT
            )
        """)
        cur.execute("""
            INSERT INTO inventory_new (item_id, category, name, price, member_price, remark)
            SELECT item_id, category, name, price, member_price, remark FROM inventory
        """)
        cur.execute("DROP TABLE inventory")
        cur.execute("ALTER TABLE inventory_new RENAME TO inventory")

    # sales 表
    cur.execute("""
        CREATE TABLE IF NOT EXISTS sales (
            sale_id INTEGER PRIMARY KEY AUTOINCREMENT,
            datetime TEXT NOT NULL,
            total_due REAL NOT NULL,
            total_paid REAL NOT NULL,
            is_member INTEGER NOT NULL DEFAULT 0
        )
    """)
    cur.execute("PRAGMA table_info(sales)")
    cols = {r[1] for r in cur.fetchall()}
    if 'member_phone' not in cols:
        cur.execute("ALTER TABLE sales ADD COLUMN member_phone TEXT")

    # sale_items 表
    cur.execute("""
        CREATE TABLE IF NOT EXISTS sale_items (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            sale_id INTEGER NOT NULL,
            category TEXT,
            name TEXT,
            price REAL,
            quantity INTEGER,
            remark TEXT,
            FOREIGN KEY(sale_id) REFERENCES sales(sale_id)
        )
    """)
    cur.execute("PRAGMA table_info(sale_items)")
    cols = {r[1] for r in cur.fetchall()}
    if 'remark' not in cols:
        cur.execute("ALTER TABLE sale_items ADD COLUMN remark TEXT")

    # memory_reminder 表
    cur.execute("""
        CREATE TABLE IF NOT EXISTS memory_reminder (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            last_reminder_date TEXT NOT NULL,
            reminder_interval INTEGER NOT NULL
        )
    """)

    # last_push_table
    cur.execute(f"""
        CREATE TABLE IF NOT EXISTS {CONFIG['database']['last_push_table']} (
            table_name TEXT PRIMARY KEY,
            last_push_time TEXT NOT NULL DEFAULT '2000-01-01 00:00:00'
        )
    """)
    for table in ["sales", "inventory", "accounting"]:
        cur.execute(f"""
            INSERT OR IGNORE INTO {CONFIG['database']['last_push_table']} 
            (table_name, last_push_time) VALUES (?, ?)
        """, (table, '2000-01-01 00:00:00'))

        # 销售目标表
        cur.execute("""
            CREATE TABLE IF NOT EXISTS sales_goals (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                period_type TEXT NOT NULL,
                target_date TEXT NOT NULL,
                target_amount REAL NOT NULL DEFAULT 0,
                created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP
            )
        """)

        # 确保每天/每月都有一个默认目标（如果不存在）
        today = date.today()
        day_key = today.strftime("%Y-%m-%d")
        month_key = today.strftime("%Y-%m")

        # 检查今日目标是否存在
        cur.execute("SELECT id FROM sales_goals WHERE period_type=? AND target_date=?",
                    ("day", day_key))
        if not cur.fetchone():
            # 若今日目标不存在，查询最近一次设置的每日目标（任意日期）
            cur.execute("""
                    SELECT target_amount FROM sales_goals 
                    WHERE period_type='day' 
                    ORDER BY target_date DESC 
                    LIMIT 1
                """)
            last_day_goal = cur.fetchone()

            if last_day_goal:
                # 存在历史每日目标，自动继承为今日目标
                cur.execute("""
                        INSERT INTO sales_goals (period_type, target_date, target_amount)
                        VALUES (?, ?, ?)
                    """, ("day", day_key, last_day_goal[0]))
            else:
                # 首次使用（无任何历史目标），可设置一个初始默认值或留空
                # 若希望完全手动设置，可删除以下代码（首次使用时目标为0，需手动设置）
                cur.execute("""
                        INSERT INTO sales_goals (period_type, target_date, target_amount)
                        VALUES (?, ?, 5000)
                    """, ("day", day_key))

        # 检查当月目标是否存在，不存在则创建默认值
        cur.execute("SELECT id FROM sales_goals WHERE period_type=? AND target_date=?",
                    ("month", month_key))
        if not cur.fetchone():
            cur.execute("""
                INSERT INTO sales_goals (period_type, target_date, target_amount)
                VALUES (?, ?, 150000)
            """, ("month", month_key))

    # 记录统计彩蛋
    cur.execute("SELECT COUNT(*) FROM sales")
    sale_count = cur.fetchone()[0]
    if sale_count >= 1000 and sale_count % 1000 == 0:
        conn.commit()
        conn.close()
        root = tk.Tk()
        root.withdraw()
        messagebox.showinfo("里程碑", f"恭喜！系统已记录第{sale_count}笔销售，再创佳绩！")
        return

    conn.commit()
    return conn

def configure_scaling_and_font(root):
    """统一设置缩放和全局字体"""
    root.tk.call('tk', 'scaling', SCALE_FACTOR)
    base_font = font.nametofont("TkDefaultFont")
    base_font.configure(family="Microsoft YaHei", size=int(10 * SCALE_FACTOR))
    style = ttk.Style()
    style.configure('.', font=(base_font.actual('family'), base_font.actual('size')))
    style.configure('Treeview', rowheight=int(24 * SCALE_FACTOR))

def make_table(parent, cols):
    """创建带滚动条的 Treeview，行高自动适配缩放"""
    container = tk.Frame(parent)
    vsb = ttk.Scrollbar(container, orient="vertical")
    vsb.pack(side='right', fill='y')
    hsb = ttk.Scrollbar(container, orient="horizontal")
    hsb.pack(side='bottom', fill='x')

    tree = ttk.Treeview(
        container,
        columns=cols,
        show='headings',
        yscrollcommand=vsb.set,
        xscrollcommand=hsb.set
    )
    vsb.config(command=tree.yview)
    hsb.config(command=tree.xview)
    tree.pack(fill='both', expand=True)
    for c in cols:
        tree.heading(c, text=c)
        tree.column(c, width=int(100 * SCALE_FACTOR))
    return tree, container


def get_version_from_filename():
    """从当前文件/EXE名称中提取版本号"""
    try:
        # 获取当前程序路径（区分脚本和EXE）
        if getattr(sys, 'frozen', False):
            # EXE打包环境
            full_path = sys.executable
        else:
            # 脚本开发环境
            full_path = os.path.abspath(__file__)

        # 提取文件名（不含路径和扩展名）
        file_name = os.path.splitext(os.path.basename(full_path))[0]

        # 正则匹配：支持 (beta|v) 后带可选空格 + 数字（可带多个小数点）
        # 兼容：beta8、beta 8、beta1.1.1、beta 8.1.1、v1、v 2.3 等
        match = re.search(r'(beta|v)\s*\d+(\.\d+)*', file_name)  # 核心修改：增加 \s* 支持可选空格
        if match:
            return match.group().replace(" ", " ")  # 移除可能的空格，统一格式（如 beta 8 → beta8）
        return "unknown"  # 匹配失败时返回默认值
    except Exception as e:
        print(f"提取版本号失败: {e}")
        return "unknown"

# 在工具函数区域添加节日检查函数
def check_festival():
    """检查当前是否为特定节日，返回节日信息"""
    today = date.today()
    # 情人节(2月14日)
    if today.month == 2 and today.day == 14:
        return {
            "name": "情人节",
            "greeting": "情人节快乐！今天销量肯定翻倍，备好货哦～",
            "title_suffix": "情人节特辑"
        }
    elif today.month == 3 and today.day == 8:
        return {
            "name": "妇女节",
            "greeting": "妇女节快乐！祝福所有妇女节日快乐哦～",
            "title_suffix": "妇女节特辑"
        }
    # 儿童节(6月1日)
    elif today.month == 6 and today.day == 1:
        return {
            "name": "儿童节",
            "greeting": "儿童节快乐！祝福所有儿童节日快乐哦～",
            "title_suffix": "儿童节特辑"
        }
    # 劳动节(5月1日)
    elif today.month == 5 and today.day == 1:
        return {
            "name": "劳动节",
            "greeting": "劳动节快乐！祝福所有劳动者们节日快乐哦～",
            "title_suffix": "劳动节特辑"
        }
    # 建党节(7月1日)
    elif today.month == 7 and today.day == 1:
        return {
            "name": "建党节",
            "greeting": "建党节快乐！祝福所有党员们节日快乐哦～",
            "title_suffix": "建党节特辑"
        }
    # 建军节(8月1日)
    elif today.month == 8 and today.day == 1:
        return {
            "name": "建军节",
            "greeting": "建军节快乐！祝福所有军人节日快乐哦～",
            "title_suffix": "建军节特辑"
        }
    # 教师节(9月1日)
    elif today.month == 9 and today.day == 10:
        return {
            "name": "教师节",
            "greeting": "教师节快乐！祝福所有教师节日快乐哦～",
            "title_suffix": "教师节特辑"
        }
    # 国庆节(10月1日)
    elif today.month == 10 and today.day == 1:
        return {
            "name": "国庆节",
            "greeting": "国庆节快乐！国庆佳节，祝祖国生日快乐！",
            "title_suffix": "国庆节特辑"
        }

    # 母亲节(5月14日)
    elif today.month == 5 and today.day == 14:
        return {
            "name": "母亲节",
            "greeting": "母亲节快乐！祝福所有母亲节日快乐哦～",
            "title_suffix": "母亲节特辑"
        }
    # 圣诞节(12月25日)
    elif today.month == 1 and today.day == 1:
        return {
            "name": "元旦",
            "greeting": "元旦快乐！祝福所有亲戚朋友节日快乐哦～",
            "title_suffix": "元旦特辑"
        }
    # 花的生日(4月27日)
    elif today.month == 4 and today.day == 27:
        return {
            "name": "生日快乐",
            "greeting": "生日快乐！祝福花生日快乐！",
            "title_suffix": "生日特辑"
        }
    # 姐的生日(3月20日)
    elif today.month == 3 and today.day == 20:
        return {
            "name": "生日快乐",
            "greeting": "生日快乐！祝福姐生日快乐日快乐哦～",
            "title_suffix": "生日特辑"
        }
    # 妹的生日(3月28日)
    elif today.month == 3 and today.day == 28:
        return {
            "name": "生日快乐",
            "greeting": "生日快乐！祝福妹生日快乐哦～",
            "title_suffix": "生日特辑"
        }

    try:
        # 1. 先创建公历对象（Solar）
        solar = Solar.fromYmd(today.year, today.month, today.day)
        # 2. 转换为农历对象（Lunar）
        lunar = solar.getLunar()
        lunar_month = lunar.getMonth()  # 农历月份
        lunar_day = lunar.getDay()  # 农历日期
    except Exception as e:
        print(f"农历转换错误: {e}")
        return None

    # 春节（正月初一）
    if lunar_month == 1 and lunar_day == 1:
        return {
            "name": "春节",
            "greeting": "春节快乐！阖家幸福，万事如意！",
            "title_suffix": "春节特辑"
        }
    # 元宵节（正月十五）
    elif lunar_month == 1 and lunar_day == 15:
        return {
            "name": "元宵节",
            "greeting": "元宵节快乐！记得吃汤圆哦～",
            "title_suffix": "元宵节特辑"
        }
    # 端午节（五月初五）
    elif lunar_month == 5 and lunar_day == 5:
        return {
            "name": "端午节",
            "greeting": "端午节快乐！祝您安康～",
            "title_suffix": "端午节特辑"
        }
    # 七夕节（七月初七）
    elif lunar_month == 7 and lunar_day == 7:
        return {
            "name": "七夕节",
            "greeting": "七夕节快乐！愿天下有情人终成眷属～",
            "title_suffix": "七夕特辑"
        }
    # 中秋节（八月十五）
    elif lunar_month == 8 and lunar_day == 15:
        return {
            "name": "中秋节",
            "greeting": "中秋节快乐！阖家团圆，幸福美满～",
            "title_suffix": "中秋节特辑"
        }
    # 重阳节（九月初九）
    elif lunar_month == 9 and lunar_day == 9:
        return {
            "name": "重阳节",
            "greeting": "重阳节快乐！祝长辈健康长寿～",
            "title_suffix": "重阳节特辑"
        }
    return None

def scale_image_to_fit(img, max_width=800, max_height=600):
    """将图片按比例缩放到最大宽高范围内"""
    width, height = img.size
    scale = min(max_width / width, max_height / height, 1.0)  # 不放大只缩小
    if scale < 1.0:
        new_width = int(width * scale)
        new_height = int(height * scale)
        return img.resize((new_width, new_height), Image.Resampling.LANCZOS)
    return img

def safe_open_image(image_path):
    """安全打开图片并返回副本，确保文件正确关闭"""
    with Image.open(image_path) as img:
        return img.copy()  # 返回副本避免文件句柄占用

# Key 文件路径（放在程序同目录）
KEY_PATH = os.path.join(os.path.dirname(__file__), 'nfc_key.bin')

def generate_key(force=False):
    if os.path.exists(KEY_PATH) and not force:
        return load_key()
    key = AESGCM.generate_key(bit_length=256)
    with open(KEY_PATH, 'wb') as f:
        f.write(key)
    try:
        os.chmod(KEY_PATH, 0o600)
    except Exception:
        pass
    return key

def load_key():
    if not os.path.exists(KEY_PATH):
        return generate_key()
    with open(KEY_PATH, 'rb') as f:
        return f.read()

def encrypt_credentials(username: str, password: str, key=None):
    """返回 dict（json 可序列化），data 为 base64(nonce + ciphertext + tag)"""
    if key is None:
        key = load_key()
    aesgcm = AESGCM(key)
    nonce = os.urandom(12)
    payload = json.dumps({"u": username, "p": password}, ensure_ascii=False).encode('utf-8')
    ct = aesgcm.encrypt(nonce, payload, None)  # ct includes tag at end
    blob = nonce + ct
    return {"v":1, "type":"login", "data": base64.b64encode(blob).decode('ascii')}

def decrypt_credentials(record: dict, key=None):
    if key is None:
        key = load_key()
    if record.get("v") != 1 or record.get("type") != "login":
        raise ValueError("Unsupported record")
    blob = base64.b64decode(record["data"].encode('ascii'))
    nonce, ct = blob[:12], blob[12:]
    aesgcm = AESGCM(key)
    payload = aesgcm.decrypt(nonce, ct, None)
    d = json.loads(payload.decode('utf-8'))
    return d["u"], d["p"]

# --- pyscard MIFARE Classic style read/write using ACR122 APDUs ---
# APDU patterns for ACR122 / PCSC readers (works in many setups)
def _send_apdu(conn, apdu):
    """发送 APDU，返回 (response_bytes, sw1, sw2)"""
    resp, sw1, sw2 = conn.transmit(apdu)
    return bytes(resp), sw1, sw2

def load_mifare_key(conn, key=b'\xff'*6, key_number=0x00):
    """Load key into reader key slot (key_number 0x00..0x01 etc)"""
    if len(key) != 6:
        raise ValueError("MIFARE key must be 6 bytes")
    apdu = [0xFF, 0x82, 0x00, key_number, 0x06] + list(key)
    _, sw1, sw2 = conn.transmit(apdu)
    return (sw1, sw2) == (0x90, 0x00)

def authenticate_block(conn, block_number, key_number=0x00, key_type=0x60):
    """
    key_type: 0x60 = Key A, 0x61 = Key B
    Use APDU: FF 86 00 00 05 01 00 <block> <key_type> <key_number>
    """
    apdu = [0xFF, 0x86, 0x00, 0x00, 0x05,
            0x01, 0x00, block_number, key_type, key_number]
    _, sw1, sw2 = conn.transmit(apdu)
    return (sw1, sw2) == (0x90, 0x00)

def read_block(conn, block_number):
    """Read 16 bytes from block. Returns bytes or raise."""
    apdu = [0xFF, 0xB0, 0x00, block_number, 0x10]
    resp, sw1, sw2 = conn.transmit(apdu)
    if (sw1, sw2) != (0x90, 0x00):
        raise RuntimeError(f"Read block {block_number} failed: SW={hex(sw1)} {hex(sw2)}")
    return bytes(resp)

# 添加销售目标相关工具函数
def get_sales_goal(period_type, target_date):
    """获取指定周期的销售目标"""
    conn = sqlite3.connect(DB_PATH)
    cur = conn.cursor()
    cur.execute("""
        SELECT target_amount FROM sales_goals 
        WHERE period_type=? AND target_date=?
    """, (period_type, target_date))
    result = cur.fetchone()
    conn.close()
    return result[0] if result else 0


def update_sales_goal(period_type, target_date, amount):
    """更新指定周期的销售目标"""
    conn = sqlite3.connect(DB_PATH)
    cur = conn.cursor()
    # 先尝试更新
    cur.execute("""
        UPDATE sales_goals SET target_amount=? 
        WHERE period_type=? AND target_date=?
    """, (amount, period_type, target_date))

    # 如果没有更新任何记录，则插入新记录
    if cur.rowcount == 0:
        cur.execute("""
            INSERT INTO sales_goals (period_type, target_date, target_amount)
            VALUES (?, ?, ?)
        """, (period_type, target_date, amount))

    conn.commit()
    conn.close()


def get_current_period_sales(period_type):
    """获取当前周期的销售总额"""
    today = date.today()
    conn = sqlite3.connect(DB_PATH)
    cur = conn.cursor()

    if period_type == "day":
        # 今日销售额
        date_str = today.strftime("%Y-%m-%d")
        cur.execute("""
            SELECT SUM(total_paid) FROM sales 
            WHERE datetime LIKE ?
        """, (f"{date_str}%",))
    else:
        # 当月销售额
        month_str = today.strftime("%Y-%m")
        cur.execute("""
            SELECT SUM(total_paid) FROM sales 
            WHERE datetime LIKE ?
        """, (f"{month_str}%",))

    result = cur.fetchone()
    conn.close()
    return result[0] if result and result[0] else 0


def get_sales_progress(period_type):
    """获取当前周期的销售进度（完成比例）"""
    today = date.today()
    target_date = today.strftime("%Y-%m-%d") if period_type == "day" else today.strftime("%Y-%m")

    goal = get_sales_goal(period_type, target_date)
    if goal <= 0:
        return 0, 0, 0  # 目标为0时，进度为0

    sales = get_current_period_sales(period_type)
    progress = min(100, (sales / goal) * 100)  # 最大100%
    return progress, sales, goal

# 全局版本号变量（程序启动时自动提取）
APP_VERSION = get_version_from_filename()

# 在全局配置区域添加单实例控制代码
class SingleInstance:
    def __init__(self):
        self.mutex_name = "SistersFlowerSystem_{8F6F0AC4-B9A1-45FD-A8CF-72F04E6BDE8F}"  # 唯一标识
        self.mutex = win32event.CreateMutex(None, False, self.mutex_name)
        self.last_error = win32api.GetLastError()

    def is_running(self):
        # 直接使用错误代码183（对应ERROR_ALREADY_EXISTS），避免依赖win32con的定义
        return self.last_error == 183

    def __del__(self):
        if hasattr(self, 'mutex') and self.mutex:
            win32api.CloseHandle(self.mutex)

class AvatarCropper:
    def __init__(self, parent, image_path):
        self.top = tk.Toplevel(parent)
        self.top.title("裁剪头像")
        self.image_path = image_path
        self.cropped_image = None  # 存储裁剪结果

        # 安全加载图片
        self.original_img = Image.open(image_path)  # 假设已实现safe_open_image
        self.base_img = self.original_img.copy()  # 保留原始尺寸用于计算
        self.scale_factor = 1.0  # 图片缩放系数
        self.offset_x = 0  # 图片偏移量（相对于画布）
        self.offset_y = 0  # 图片偏移量（相对于画布）

        # 画布尺寸
        self.canvas_width = 800
        self.canvas_height = 600
        self.canvas = tk.Canvas(
            self.top,
            width=self.canvas_width,
            height=self.canvas_height,
            bg="#f0f0f0"
        )
        self.canvas.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

        # 固定取景框（居中显示，大小200x200，可根据需求调整）
        self.crop_size = 200
        self.crop_x1 = (self.canvas_width - self.crop_size) // 2
        self.crop_y1 = (self.canvas_height - self.crop_size) // 2
        self.crop_x2 = self.crop_x1 + self.crop_size
        self.crop_y2 = self.crop_y1 + self.crop_size

        # 创建固定取景框
        self.crop_rect = self.canvas.create_rectangle(
            self.crop_x1, self.crop_y1, self.crop_x2, self.crop_y2,
            outline="red", width=2, dash=(5, 2)
        )

        # 初始显示图片（居中）
        self.initial_image_position()
        self.update_display_image()

        # 绘制遮罩（固定在取景框周围）
        self.mask_items = []
        self.draw_mask()

        # 绑定事件（改为拖动图片，而非取景框）
        self.canvas.bind("<Button-1>", self.start_drag_image)
        self.canvas.bind("<B1-Motion>", self.drag_image)
        self.canvas.bind("<MouseWheel>", self.zoom_image)  # 滚轮缩放图片
        self.canvas.bind("<Button-3>", self.reset_view)  # 右键重置视图

        # 按钮
        btn_frame = tk.Frame(self.top)
        btn_frame.pack(pady=10)
        ttk.Button(btn_frame, text="确认裁剪", command=self.crop).pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame, text="取消", command=self.top.destroy).pack(side=tk.LEFT, padx=5)

        # 状态变量
        self.dragging_image = False
        self.start_drag_x = 0
        self.start_drag_y = 0

    def initial_image_position(self):
        """初始化图片位置（居中显示）"""
        # 计算初始缩放系数（确保图片能完整显示在画布中）
        img_width, img_height = self.base_img.size
        scale_w = self.canvas_width / img_width
        scale_h = self.canvas_height / img_height
        self.scale_factor = min(scale_w, scale_h) * 0.9  # 留10%边距

        # 初始偏移量（居中）
        self.offset_x = self.canvas_width // 2
        self.offset_y = self.canvas_height // 2

    def update_display_image(self):
        """更新显示的图片（考虑缩放和偏移）"""
        # 计算缩放后的图片尺寸
        img_width, img_height = self.base_img.size
        scaled_width = int(img_width * self.scale_factor)
        scaled_height = int(img_height * self.scale_factor)

        # 缩放图片
        self.scaled_img = self.base_img.resize(
            (scaled_width, scaled_height),
            Image.Resampling.LANCZOS
        )

        # 转换为Tkinter可用格式
        self.tk_image = ImageTk.PhotoImage(self.scaled_img)

        # 更新画布图片
        if hasattr(self, 'image_id'):
            self.canvas.delete(self.image_id)
        # 图片锚点为中心，基于偏移量定位
        self.image_id = self.canvas.create_image(
            self.offset_x, self.offset_y,
            image=self.tk_image,
            anchor=tk.CENTER
        )
        self.canvas.tag_lower(self.image_id)  # 确保图片在底层
        self.canvas.tag_raise(self.crop_rect)  # 确保取景框在顶层

    def draw_mask(self):
        """绘制取景框外的遮罩（固定位置）"""
        # 清除旧遮罩
        for item in self.mask_items:
            self.canvas.delete(item)
        self.mask_items = []

        # 遮罩颜色（半透明灰色）
        mask_color = "#888888"
        # 绘制取景框周围的四个遮罩区域
        self.mask_items.extend([
            # 左方遮罩
            self.canvas.create_rectangle(
                0, 0, self.crop_x1, self.canvas_height,
                fill=mask_color, outline="", stipple="gray50"
            ),
            # 右方遮罩
            self.canvas.create_rectangle(
                self.crop_x2, 0, self.canvas_width, self.canvas_height,
                fill=mask_color, outline="", stipple="gray50"
            ),
            # 上方遮罩
            self.canvas.create_rectangle(
                self.crop_x1, 0, self.crop_x2, self.crop_y1,
                fill=mask_color, outline="", stipple="gray50"
            ),
            # 下方遮罩
            self.canvas.create_rectangle(
                self.crop_x1, self.crop_y2, self.crop_x2, self.canvas_height,
                fill=mask_color, outline="", stipple="gray50"
            )
        ])
        self.canvas.tag_raise(self.crop_rect)  # 确保取景框在顶层

    def start_drag_image(self, event):
        """开始拖拽图片"""
        # 检查点击位置是否在图片上
        img_width, img_height = self.scaled_img.size
        img_x1 = self.offset_x - img_width // 2
        img_y1 = self.offset_y - img_height // 2
        img_x2 = self.offset_x + img_width // 2
        img_y2 = self.offset_y + img_height // 2

        if img_x1 <= event.x <= img_x2 and img_y1 <= event.y <= img_y2:
            self.dragging_image = True
            self.start_drag_x = event.x
            self.start_drag_y = event.y

    def drag_image(self, event):
        """拖拽图片（移动图片位置）"""
        if not self.dragging_image:
            return

        # 计算移动距离
        dx = event.x - self.start_drag_x
        dy = event.y - self.start_drag_y

        # 更新图片偏移量
        self.offset_x += dx
        self.offset_y += dy

        # 更新显示
        self.update_display_image()
        self.start_drag_x = event.x
        self.start_drag_y = event.y

    def zoom_image(self, event):
        """缩放图片（滚轮控制）"""
        # 记录当前鼠标在画布上的位置
        mouse_x, mouse_y = event.x, event.y

        # 计算新缩放系数
        old_scale = self.scale_factor
        if event.delta > 0:
            self.scale_factor = min(5.0, self.scale_factor * 1.1)  # 最大放大5倍
        else:
            self.scale_factor = max(0.2, self.scale_factor / 1.1)  # 最小缩小到0.2倍

        # 以鼠标位置为中心缩放（保持鼠标指向的图片位置不变）
        # 计算公式：新偏移量 = 鼠标位置 - (鼠标位置 - 旧偏移量) * (新缩放 / 旧缩放)
        self.offset_x = mouse_x - (mouse_x - self.offset_x) * (self.scale_factor / old_scale)
        self.offset_y = mouse_y - (mouse_y - self.offset_y) * (self.scale_factor / old_scale)

        # 更新显示
        self.update_display_image()

    def reset_view(self, event):
        """重置视图（右键点击）"""
        self.initial_image_position()
        self.update_display_image()
        self.draw_mask()

    def crop(self):
        """执行裁剪（计算固定取景框对应的图片区域）"""
        # 1. 获取图片在画布上的显示区域（缩放后的位置和尺寸）
        scaled_width = int(self.base_img.width * self.scale_factor)
        scaled_height = int(self.base_img.height * self.scale_factor)
        img_display_x1 = self.offset_x - scaled_width // 2
        img_display_y1 = self.offset_y - scaled_height // 2

        # 2. 计算固定取景框与图片显示区域的重叠部分
        overlap_x1 = max(self.crop_x1, img_display_x1)
        overlap_y1 = max(self.crop_y1, img_display_y1)
        overlap_x2 = min(self.crop_x2, img_display_x1 + scaled_width)
        overlap_y2 = min(self.crop_y2, img_display_y1 + scaled_height)

        # 3. 检查是否有重叠区域
        if overlap_x1 >= overlap_x2 or overlap_y1 >= overlap_y2:
            messagebox.showerror("错误", "请确保取景框内有图片内容")
            return

        # 4. 将重叠区域转换为缩放后图片上的坐标（相对图片左上角）
        img_rel_x1 = overlap_x1 - img_display_x1
        img_rel_y1 = overlap_y1 - img_display_y1
        img_rel_x2 = overlap_x2 - img_display_x1
        img_rel_y2 = overlap_y2 - img_display_y1

        # 5. 将缩放后图片的坐标转换为原始图片的坐标
        orig_x1 = int(img_rel_x1 / self.scale_factor)
        orig_y1 = int(img_rel_y1 / self.scale_factor)
        orig_x2 = int(img_rel_x2 / self.scale_factor)
        orig_y2 = int(img_rel_y2 / self.scale_factor)

        # 6. 执行裁剪
        self.cropped_image = self.original_img.crop((orig_x1, orig_y1, orig_x2, orig_y2))

        # 释放资源并关闭窗口
        self.original_img.close()
        self.top.destroy()

class WindowsNotification:
    def __init__(self, title, message, icon_path=None, duration=10):
        self.title = title
        self.message = message
        self.icon_path = icon_path
        self.duration = duration * 1000  # 转换为毫秒
        self.hwnd = None
        self.notify_id = 1001  # 通知唯一标识
        self.timer_id = None  # 定时器ID
        self.class_atom = None  # 新增：保存窗口类原子值
        # 生成唯一类名（使用线程ID确保唯一性）
        self.class_name = f"NotificationClass_{threading.get_ident()}"

    def _wnd_proc(self, hwnd, msg, wparam, lparam):
        """窗口消息处理函数"""
        if msg == win32con.WM_DESTROY:
            win32gui.PostQuitMessage(0)
            return 0
        elif msg == win32con.WM_TIMER:
            # 处理定时器事件
            self._on_timer()
            return 0
        elif msg == win32con.WM_COMMAND:
            return 0
        # 其他消息使用默认处理
        return win32gui.DefWindowProc(hwnd, msg, wparam, lparam)

    def _create_window(self):
        """创建隐藏窗口用于处理通知消息"""
        wc = win32gui.WNDCLASS()
        wc.lpfnWndProc = self._wnd_proc
        wc.hInstance = win32api.GetModuleHandle(None)
        wc.lpszClassName = self.class_name  # 使用唯一类名
        self.class_atom = win32gui.RegisterClass(wc)  # 保存类原子
        self.hwnd = win32gui.CreateWindow(
            self.class_atom, "Notification Window", 0, 0, 0,
            win32con.CW_USEDEFAULT, win32con.CW_USEDEFAULT,
            0, 0, wc.hInstance, None
        )
        win32gui.UpdateWindow(self.hwnd)

    def _load_icon(self):
        """加载通知图标"""
        if self.icon_path and os.path.exists(self.icon_path):
            try:
                return win32gui.LoadImage(
                    win32api.GetModuleHandle(None),
                    self.icon_path,
                    win32con.IMAGE_ICON,
                    0, 0,
                    win32con.LR_LOADFROMFILE | win32con.LR_DEFAULTSIZE
                )
            except:
                pass
        # 加载默认图标
        return win32gui.LoadIcon(0, win32con.IDI_APPLICATION)

    # 修改WindowsNotification类中的show()方法里的消息循环部分
    def show(self):
        """显示系统通知"""
        try:  # 增加try-finally确保资源释放
            self._create_window()
            hicon = self._load_icon()

            # 初始化通知图标
            nid = (
                self.hwnd, self.notify_id,
                win32gui.NIF_ICON | win32gui.NIF_MESSAGE | win32gui.NIF_TIP,
                win32con.WM_USER + 20, hicon, "姐妹花销售系统"
            )
            win32gui.Shell_NotifyIcon(win32gui.NIM_ADD, nid)

            # 更新通知内容
            nid = (
                self.hwnd, self.notify_id,
                win32gui.NIF_INFO,
                win32con.WM_USER + 20, hicon,
                "通知", self.message, 200, self.title
            )
            win32gui.Shell_NotifyIcon(win32gui.NIM_MODIFY, nid)

            # 调用底层 SetTimer API
            self.timer_id = user32.SetTimer(
                self.hwnd,  # 窗口句柄
                0,  # 定时器ID（0表示自动分配）
                self.duration,  # 超时时间（毫秒）
                None  # 回调函数（通过WM_TIMER消息处理）
            )
            if not self.timer_id:
                raise ctypes.WinError(ctypes.get_last_error())

            # 消息循环
            while True:
                # GetMessage实际返回值为 (msg, wparam, lparam) 或 (False,) 当无消息时
                msg_result = win32gui.GetMessage(None, 0, 0)

                # 检查是否有消息（返回值长度为3表示有消息）
                if len(msg_result) != 3:
                    break  # 无消息时退出循环

                msg, wparam, lparam = msg_result

                # 传递消息给翻译和分发函数
                win32gui.TranslateMessage((msg, wparam, lparam))
                win32gui.DispatchMessage((msg, wparam, lparam))

                # 检查退出消息
                if msg == win32con.WM_QUIT:
                    break
        finally:
            # 清理资源
            if self.timer_id:
                user32.KillTimer(self.hwnd, self.timer_id)
            # 删除通知图标
            win32gui.Shell_NotifyIcon(win32gui.NIM_DELETE, (self.hwnd, self.notify_id))
            # 销毁窗口
            if self.hwnd:
                win32gui.DestroyWindow(self.hwnd)
            # 注销窗口类（关键修复）
            if self.class_atom:
                try:
                    win32gui.UnregisterClass(self.class_atom, win32api.GetModuleHandle(None))
                except:
                    pass

class LoginWindow:
    def __init__(self, root):
        self.master = root
        self.master.title("登录")
        self.master.resizable(False, False)
        self.style = ttk.Style()

        # 应用缩放和字体
        configure_scaling_and_font(self.master)

        # 登录框架
        self.frame = tk.Frame(self.master)
        self.frame.pack(fill='both', expand=True, padx=20, pady=20)

        # 新增头像显示区域
        self.avatar_frame = tk.Frame(self.frame)
        self.avatar_frame.grid(row=0, column=2, rowspan=3, padx=20)

        self.avatar_img = tk.Label(self.avatar_frame)
        self.avatar_img.pack(pady=10)

        # 用户名
        tk.Label(self.frame, text="用户名：").grid(row=0, column=0, sticky='e', pady=5)
        cursor_conn = sqlite3.connect(DB_PATH, check_same_thread=False)
        cursor_cur = cursor_conn.cursor()
        cursor_cur.execute("SELECT username FROM users")
        users = [r[0] for r in cursor_cur.fetchall()]
        cursor_conn.close()

        self.username_var = tk.StringVar()
        self.user_cb = ttk.Combobox(self.frame, textvariable=self.username_var, values=users, state='normal', width=25)
        self.user_cb.grid(row=0, column=1, pady=5)
        self.user_cb.focus()

        # 当切换用户时更新头像（绑定事件）
        try:
            self.user_cb.bind("<<ComboboxSelected>>", self.update_avatar)
        except Exception:
            pass

        # 密码框和登录按钮组合
        tk.Label(self.frame, text="密 码：").grid(row=1, column=0, sticky='e', pady=5)

        # 创建容器放置密码框和登录按钮
        self.pwd_frame = ttk.Frame(self.frame)
        self.pwd_frame.grid(row=1, column=1, pady=5, sticky='w')

        # 缩短密码框，留出按钮空间
        self.password_var = tk.StringVar()
        self.pwd_entry = ttk.Entry(self.pwd_frame, textvariable=self.password_var, show='*', width=20)
        self.pwd_entry.pack(side=tk.LEFT, padx=(0, 2))

        # 登录按钮改为向左箭头，与输入框组合后总长度和用户名框一致
        self.login_btn = ttk.Button(self.pwd_frame, text="←", command=self.login, width=3)
        self.login_btn.pack(side=tk.LEFT)

        # 显示密码复选框
        self.show_pwd_var = tk.BooleanVar(value=False)
        self.show_pwd_check = ttk.Checkbutton(
            self.frame,
            text="显示密码",
            variable=self.show_pwd_var,
            command=self.toggle_password_visibility
        )
        self.show_pwd_check.grid(row=2, column=1, sticky='w', pady=5)

        # 绑定回车
        self.user_cb.bind("<Return>", lambda e: self.pwd_entry.focus())
        self.pwd_entry.bind("<Return>", lambda e: self.login())

        # 初始显示第一个用户的头像
        if users:
            self.username_var.set(users[0])
            self.update_avatar(None)

    def update_avatar(self, event):
        """更新显示选中用户的头像"""
        username = self.username_var.get()
        if not username:
            return

        try:
            conn = sqlite3.connect(DB_PATH, check_same_thread=False)
            cur = conn.cursor()
            cur.execute("SELECT avatar FROM users WHERE username=?", (username,))
            result = cur.fetchone()
            conn.close()

            if result and result[0]:
                avatar_path = result[0]
            else:
                avatar_path = "profile photo.png"  # 默认头像

            # 检查文件是否存在
            if not os.path.exists(avatar_path):
                # resource_path 是你程序里用于查找资源的函数（保持原有逻辑）
                try:
                    avatar_path = resource_path("profile photo.png")
                except Exception:
                    avatar_path = None

            # 加载并显示头像
            if avatar_path and os.path.exists(avatar_path):
                img = Image.open(avatar_path)
                img.thumbnail((150, 150))  # 缩放头像
                photo = ImageTk.PhotoImage(img)
                self.avatar_img.config(image=photo, text="")
                self.avatar_img.image = photo  # 保持引用
            else:
                # 显示默认占位
                default_img = tk.PhotoImage(width=150, height=150)
                self.avatar_img.config(image=default_img, text="无头像")
                self.avatar_img.image = default_img

        except Exception as e:
            print(f"加载头像失败: {e}")
            # 显示默认头像占位
            default_img = tk.PhotoImage(width=150, height=150)
            self.avatar_img.config(image=default_img, text="无头像")
            self.avatar_img.image = default_img

    def toggle_password_visibility(self):
        """切换密码显示状态"""
        if self.show_pwd_var.get():
            # 显示密码
            self.pwd_entry.config(show='')
        else:
            # 隐藏密码
            self.pwd_entry.config(show='*')

    def check_memory_reminder(self):
        try:
            conn = sqlite3.connect(DB_PATH, check_same_thread=False)
            cur = conn.cursor()

            # 获取提醒设置
            cur.execute("SELECT last_reminder_date, reminder_interval FROM memory_reminder")
            row = cur.fetchone()

            if not row:
                conn.close()
                return  # 未设置提醒

            last_date_str, interval = row
            last_date = datetime.strptime(last_date_str, "%Y-%m-%d")
            current_date = datetime.now()

            # 计算月份差
            months_diff = (current_date.year - last_date.year) * 12 + (current_date.month - last_date.month)

            if months_diff >= interval:
                # 需要提醒，更新最后提醒日期
                now_str = current_date.strftime("%Y-%m-%d")
                cur.execute("UPDATE memory_reminder SET last_reminder_date=?", (now_str,))
                conn.commit()

                # 获取文件统计信息（修改为检查指定文件和文件夹）
                stats = self.get_database_stats()

                # 显示提醒
                self.show_memory_reminder(stats)

            conn.close()
        except Exception as e:
            print(f"内存提醒检查失败: {e}")

    def get_database_stats(self):
        try:
            # 获取根目录路径（假设当前工作目录为根目录）
            root_path = os.getcwd()

            # 1. 计算backups文件夹大小
            backups_size = self.get_folder_size(os.path.join(root_path, "backups"))

            # 2. 计算SistersOS_log文件夹大小
            log_size = self.get_folder_size(os.path.join(root_path, "SistersOS_log"))

            # 3. 计算数据库文件大小
            db_path = os.path.join(root_path, "sisters_flowers.db")
            db_size = os.path.getsize(db_path) if os.path.exists(db_path) else 0

            # 4. 计算根目录总大小
            root_total_size = self.get_folder_size(root_path)

            # 5. 计算其他文件大小
            other_size = root_total_size - backups_size - log_size - db_size

            return {
                "backups_size": backups_size,
                "log_size": log_size,
                "db_size": db_size,
                "other_size": other_size,
                "root_total_size": root_total_size
            }
        except Exception as e:
            print(f"获取文件统计信息失败: {e}")
            return None

    def get_folder_size(self, folder_path):
        """计算文件夹总大小（包括子文件夹）"""
        total_size = 0
        if not os.path.exists(folder_path) or not os.path.isdir(folder_path):
            return total_size

        try:
            for item in os.listdir(folder_path):
                item_path = os.path.join(folder_path, item)
                if os.path.isfile(item_path):
                    total_size += os.path.getsize(item_path)
                elif os.path.isdir(item_path):
                    total_size += self.get_folder_size(item_path)
        except Exception as e:
            print(f"计算文件夹 {folder_path} 大小时出错: {e}")
        return total_size

    def format_size(self, size_bytes):
        """将字节转换为合适的单位表示"""
        units = ['B', 'KB', 'MB', 'GB']
        unit_index = 0

        while size_bytes >= 1024 and unit_index < len(units) - 1:
            size_bytes /= 1024
            unit_index += 1

        return f"{size_bytes:.2f} {units[unit_index]}"

    def show_memory_reminder(self, stats):
        if not stats:
            return

        message = (
            f"存储空间统计信息:\n\n"
            f"备份文件夹: {self.format_size(stats['backups_size'])}\n"
            f"日志文件夹: {self.format_size(stats['log_size'])}\n"
            f"数据库文件: {self.format_size(stats['db_size'])}\n"
            f"其他文件: {self.format_size(stats['other_size'])}\n\n"
            f"根目录总大小: {self.format_size(stats['root_total_size'])}"
        )

        messagebox.showinfo("存储空间提醒", message)

    def login(self, u=None, p=None):
        # 支持传入用户名/密码（用于 NFC 自动登录），也支持无参数由界面输入
        if u is None and p is None:
            u = self.username_var.get().strip()
            p = self.password_var.get().strip()
        else:
            # 如果是从 NFC 或外部调用，更新界面输入框显示（可选）
            try:
                self.username_var.set(u)
                self.password_var.set(p)
            except Exception:
                pass

        if not u or not p:
            return messagebox.showerror("错误", "用户名和密码不能为空")
        try:
            conn = sqlite3.connect(DB_PATH, check_same_thread=False)
            cur = conn.cursor()
            cur.execute("SELECT password FROM users WHERE username=?", (u,))
            row = cur.fetchone()
            conn.close()
        except Exception as e:
            return messagebox.showerror("错误", f"数据库访问失败：{e}")

        if not row or row[0] != p:
            return messagebox.showerror("登录失败", "用户名或密码错误")

        # 关键修改：密码登录成功，标记已登录状态（终止NFC自动登录）
        self._is_logged_in = True  # ↑↑↑ 新增：设置为已登录 ↑↑↑

        # 登陆成功，提示余额低于5元的会员
        try:
            conn = sqlite3.connect(DB_PATH, check_same_thread=False)
            cur = conn.cursor()
            cur.execute("SELECT phone, balance FROM members WHERE balance < 5")
            low = cur.fetchall()
            conn.close()
            if low:
                msg = "\n".join(f"{ph}: {bal:.2f}元" for ph, bal in low)
                messagebox.showinfo("余额提醒", f"以下会员余额低于5元：\n{msg}")
        except:
            pass  # 提示失败不影响登录

        # 切换到主界面
        self.frame.destroy()
        # 检查是否需要内存提醒
        self.check_memory_reminder()
        MainApp(self.master, u)

class MainApp:
    def __init__(self, root, username):
        self.root = root
        self.username = username
        self.root.title(f"姐妹花销售系统 {APP_VERSION}")
        self.root.resizable(True, True)

        # 添加系统托盘相关初始化
        self.tray_icon = None
        self.tray_running = False  # 自定义变量跟踪托盘状态
        self.tray_lock = threading.Lock()  # 新增：线程锁，防止并发操作
        self.create_tray_icon()

        # 绑定窗口关闭事件
        self.root.protocol("WM_DELETE_WINDOW", self.on_close_request)

        # 应用缩放与字体（再次确保）
        configure_scaling_and_font(root)

        # Notebook 选项卡
        notebook = ttk.Notebook(root)
        notebook.pack(fill='both', expand=True)

        # 创建各页面
        self.sales_page = SalesPage(notebook, self)
        self.members_page = MemberPage(notebook)
        self.account_page = AccountPage(notebook, self)
        self.inventory_page = InventoryPage(notebook)
        self.settings_page = SettingsPage(notebook, self)

        # 添加到 Notebook
        notebook.add(self.sales_page.frame, text="销售")
        notebook.add(self.members_page.frame, text="会员")
        notebook.add(self.account_page.frame, text="记账")
        notebook.add(self.inventory_page.frame, text="库存")
        notebook.add(self.settings_page.frame, text="设置")

        # 底部状态栏
        status = ttk.Frame(root)
        status.pack(side=tk.BOTTOM, fill=tk.X)

        self.user_label = ttk.Label(status, text=f"当前用户：{self.username}")
        self.user_label.pack(side=tk.LEFT, padx=5)

        self.time_label = ttk.Label(status, text="")
        self.time_label.pack(side=tk.RIGHT, padx=5)

        # 添加销售目标进度展示区域
        self.progress_frame = tk.Frame(root)
        self.progress_frame.pack(fill=tk.X, padx=10, pady=5)

        # 每日进度
        self.day_progress_frame = tk.Frame(self.progress_frame)
        self.day_progress_frame.pack(fill=tk.X, pady=2)

        ttk.Label(self.day_progress_frame, text="今日销售目标:").pack(side=tk.LEFT, padx=5)
        self.day_progress_var = tk.DoubleVar()
        self.day_progress_bar = ttk.Progressbar(
            self.day_progress_frame,
            variable=self.day_progress_var,
            maximum=100,
            length=300
        )
        self.day_progress_bar.pack(side=tk.LEFT, padx=5, fill=tk.X, expand=True)
        self.day_progress_label = ttk.Label(self.day_progress_frame, text="0% (¥0/¥0)")
        self.day_progress_label.pack(side=tk.LEFT, padx=5)

        # 每月进度
        self.month_progress_frame = tk.Frame(self.progress_frame)
        self.month_progress_frame.pack(fill=tk.X, pady=2)

        ttk.Label(self.month_progress_frame, text="本月销售目标:").pack(side=tk.LEFT, padx=5)
        self.month_progress_var = tk.DoubleVar()
        self.month_progress_bar = ttk.Progressbar(
            self.month_progress_frame,
            variable=self.month_progress_var,
            maximum=100,
            length=300
        )
        self.month_progress_bar.pack(side=tk.LEFT, padx=5, fill=tk.X, expand=True)
        self.month_progress_label = ttk.Label(self.month_progress_frame, text="0% (¥0/¥0)")
        self.month_progress_label.pack(side=tk.LEFT, padx=5)

        # 启动时钟
        self._running = True
        self.update_clock()

        # 从配置读取推送参数（替换硬编码）
        self.push_interval = int(CONFIG["push"]["interval"])
        self.mobile_endpoint = CONFIG["push"]["endpoint"]

        # 启动推送线程
        self.start_test_server()
        self._stop_push = threading.Event()
        threading.Thread(target=self._push_loop, daemon=True).start()
        # 启动备份线程
        self.start_backup_scheduler()
        self.root.bind("<Control-Shift-D>", self.show_developer_info)
        # 添加节日彩蛋检查
        festival = check_festival()
        if festival:
            # 显示节日祝福弹窗
            messagebox.showinfo(f"{festival['name']}快乐", festival['greeting'])
            # 修改窗口标题为节日主题
            self.root.title(f"姐妹花销售系统 {APP_VERSION} - {festival['title_suffix']}")
        # 初始化通知支持
        self.notification_supported = self.check_notification_support()
        self.game_click_count = 0  # 初始化点击计数器
        self.tab_switch_count = 0
        self.update_sales_progress()
        self.root.bind("<<NotebookTabChanged>>", self.count_tab_switch)

    def update_sales_progress(self):
        """更新销售进度展示"""
        # 更新每日进度
        day_progress, day_sales, day_goal = get_sales_progress("day")
        self.day_progress_var.set(day_progress)
        self.day_progress_label.config(
            text=f"{day_progress:.1f}% (¥{day_sales:.2f}/¥{day_goal:.2f})"
        )

        # 根据进度设置不同颜色
        if day_progress >= 100:
            self.day_progress_bar.config(style="success.Horizontal.TProgressbar")
            # 检查是否是刚刚达成目标，若是则显示庆祝消息
            if not hasattr(self, 'day_goal_achieved'):
                self.show_goal_achievement("今日销售目标达成！")
                self.day_goal_achieved = True
        elif day_progress >= 80:
            self.day_progress_bar.config(style="warning.Horizontal.TProgressbar")
        else:
            self.day_progress_bar.config(style="info.Horizontal.TProgressbar")

        # 更新每月进度
        month_progress, month_sales, month_goal = get_sales_progress("month")
        self.month_progress_var.set(month_progress)
        self.month_progress_label.config(
            text=f"{month_progress:.1f}% (¥{month_sales:.2f}/¥{month_goal:.2f})"
        )

        if month_progress >= 100:
            self.month_progress_bar.config(style="success.Horizontal.TProgressbar")
            if not hasattr(self, 'month_goal_achieved'):
                self.show_goal_achievement("本月销售目标达成！")
                self.month_goal_achieved = True
        elif month_progress >= 80:
            self.month_progress_bar.config(style="warning.Horizontal.TProgressbar")
        else:
            self.month_progress_bar.config(style="info.Horizontal.TProgressbar")

        # 每1分钟更新一次
        if self._running:
            self.root.after(60000, self.update_sales_progress)

    def show_goal_achievement(self, message):
        """显示目标达成庆祝消息"""
        # 显示弹窗提示
        messagebox.showinfo("目标达成！", f"{message}\n恭喜团队取得优异成绩！")

        # 同时发送系统通知
        self.show_system_notification("销售目标达成", message)

    def count_tab_switch(self, event):
        self.tab_switch_count += 1
        if self.tab_switch_count == 10:
            messagebox.showinfo("成就解锁", "获得「标签页冲浪者」称号！\n你似乎对切换标签有特殊执念~")
        if self.tab_switch_count == 20:
            messagebox.showinfo("成就解锁", "获得「标签页流浪者者」称号！\n你似乎喜欢在标签页游走~")
        if self.tab_switch_count == 30:
            messagebox.showinfo("闲的了", "那你点吧")

    def check_notification_support(self):
        """检查当前系统是否支持通知"""
        if platform.system() != "Windows":
            return False
        try:
            version = platform.release()
            # 支持Windows 8.1 (6.3)、Windows 10 (10.0) 和 Windows 11
            return version in ["8.1", "10", "11"] or version == "6.3"
        except:
            return False

    def show_system_notification(self, title, message):
        """显示系统通知"""
        if not self.notification_supported:
            return False

        try:
            # 获取图标路径
            icon_path = resource_path("f.ico")
            # 在独立线程中显示通知，避免阻塞主程序
            threading.Thread(
                target=lambda: WindowsNotification(title, message, icon_path).show(),
                daemon=True
            ).start()
            return True
        except Exception as e:
            print(f"显示通知失败: {e}")
            return False

    def show_developer_info(self, event):
        """显示自定义开发者信息窗口（修改版，添加游戏触发）"""
        dev_window = tk.Toplevel(self.root)
        dev_window.title("关于开发者")
        dev_window.geometry("300x200")
        dev_window.resizable(False, False)
        dev_window.transient(self.root)
        dev_window.grab_set()

        # 居中显示
        dev_window.update_idletasks()
        width = dev_window.winfo_width()
        height = dev_window.winfo_height()
        x = (self.root.winfo_width() // 2) - (width // 2) + self.root.winfo_x()
        y = (self.root.winfo_height() // 2) - (height // 2) + self.root.winfo_y()
        dev_window.geometry(f"+{x}+{y}")

        # 开发者信息列表
        info = [
            "软件名称：SistersOS",
            "开发者：吴小零",
            "联系方式：15199961779",
            "开发日期：2025年5月28日",
            "感谢使用本系统！"  # 这行将作为触发点
        ]

        for i, line in enumerate(info):
            if "联系方式" in line:
                contact_label = ttk.Label(dev_window, text=line, foreground="blue", cursor="hand2")
                contact_label.grid(row=i, column=0, sticky="w", padx=20, pady=5)
                contact_label.bind("<Double-1>", lambda e: self.show_minicalculator(dev_window))
            else:
                # 为最后一行添加点击事件
                if i == len(info) - 1:
                    thanks_label = ttk.Label(dev_window, text=line, cursor="hand2")
                    thanks_label.grid(row=i, column=0, sticky="w", padx=20, pady=5)
                    # 绑定点击事件到游戏触发函数
                    thanks_label.bind("<Button-1>", lambda e, win=dev_window: self.trigger_whack_a_mole(win))
                else:
                    ttk.Label(dev_window, text=line).grid(row=i, column=0, sticky="w", padx=20, pady=5)

    def trigger_whack_a_mole(self, parent_window):
        """计数点击次数，达到3次触发游戏"""
        self.game_click_count += 1
        if self.game_click_count >= 3:
            self.game_click_count = 0  # 重置计数器
            parent_window.destroy()  # 关闭开发者窗口
            self.show_whack_a_mole()  # 显示游戏窗口

    def show_whack_a_mole(self):
        """显示打地鼠游戏窗口"""
        game_window = tk.Toplevel(self.root)
        game_window.title("像素打地鼠 🐭")
        game_window.geometry("300x300")
        game_window.resizable(False, False)
        game_window.transient(self.root)
        game_window.grab_set()

        # 游戏变量
        score = 0
        time_left = 30  # 30秒游戏时间
        mole_pos = None
        game_running = True

        # 分数显示
        score_var = tk.StringVar(value=f"分数: 0")
        ttk.Label(game_window, textvariable=score_var, font=("SimHei", 14)).pack(pady=10)

        # 时间显示
        time_var = tk.StringVar(value=f"时间: {time_left}秒")
        ttk.Label(game_window, textvariable=time_var, font=("SimHei", 14)).pack(pady=5)

        # 游戏网格（3x3地鼠洞）
        grid_frame = ttk.Frame(game_window)
        grid_frame.pack(pady=20)
        cells = []
        for i in range(3):
            row = []
            for j in range(3):
                cell = ttk.Label(
                    grid_frame,
                    text="⬜",  # 初始显示空白格子
                    font=("SimHei", 30),
                    width=2,
                    anchor="center"
                )
                cell.grid(row=i, column=j, padx=5, pady=5)
                row.append(cell)
            cells.append(row)

        def show_mole():
            """随机显示地鼠"""
            nonlocal mole_pos
            # 隐藏之前的地鼠
            if mole_pos:
                i, j = mole_pos
                cells[i][j].config(text="⬜")

            # 随机位置
            if game_running:
                i = random.randint(0, 2)
                j = random.randint(0, 2)
                cells[i][j].config(text="🐭")  # 地鼠图标
                mole_pos = (i, j)
                # 随机时间后隐藏（500-1000毫秒）
                game_window.after(random.randint(500, 1000), show_mole)

        def whack(event):
            """点击地鼠计分"""
            nonlocal score, mole_pos
            if mole_pos:
                i, j = mole_pos
                if event.widget == cells[i][j]:
                    score += 10
                    score_var.set(f"分数: {score}")
                    cells[i][j].config(text="💥")  # 击中效果
                    # 记录当前位置用于延迟恢复
                    current_cell = cells[i][j]
                    # 500毫秒后恢复为白格子
                    game_window.after(500, lambda: current_cell.config(text="⬜"))
                    mole_pos = None  # 清除当前位置

        # 绑定所有格子的点击事件
        for i in range(3):
            for j in range(3):
                cells[i][j].bind("<Button-1>", whack)

        def update_timer():
            """更新游戏时间"""
            nonlocal time_left, game_running
            if time_left > 0 and game_running:
                time_left -= 1
                time_var.set(f"时间: {time_left}秒")
                game_window.after(1000, update_timer)
            elif time_left == 0:
                end_game()

        def end_game():
            """结束游戏并显示评语"""
            nonlocal game_running
            game_running = False
            # 隐藏所有地鼠
            for i in range(3):
                for j in range(3):
                    cells[i][j].config(text="⬜")

            # 根据分数显示评语
            if score >= 350:
                comment = "老实交代，是不是用连点器了？"
            elif score >= 200:
                comment = "手速超越99%凉皮店老板！"
            elif score >= 100:
                comment = "不错哦，再练练能赶上松鼠啦！"
            elif score >= 50:
                comment = "还行还行，至少没打空太多～"
            else:
                comment = "慢慢来，地鼠都睡著了～"

            # 显示结果
            result = f"游戏结束！\n最终得分: {score}\n{comment}"
            messagebox.showinfo("游戏结束", result, parent=game_window)
            game_window.destroy()

        # 开始游戏
        import random  # 确保导入随机模块
        show_mole()
        update_timer()

        # 窗口关闭时终止游戏
        game_window.protocol("WM_DELETE_WINDOW",
                             lambda: setattr(locals(), 'game_running', False) or game_window.destroy())

    def show_minicalculator(self, parent):
        """显示迷你计算器"""
        calc = tk.Toplevel(parent)
        calc.title("迷你计算器")
        calc.geometry("300x300")
        calc.resizable(False, False)
        calc.transient(parent)

        # 计算器显示区域
        display_var = tk.StringVar()
        display = ttk.Entry(calc, textvariable=display_var, justify="right", font=("Arial", 14))
        display.grid(row=0, column=0, columnspan=4, padx=5, pady=5, sticky="nsew")
        display.config(state="readonly")

        # 按钮点击处理
        def button_click(value):
            current = display_var.get()
            if value == "C":
                display_var.set("")
            elif value == "=":
                try:
                    # 支持基本运算
                    result = eval(current.replace("×", "*").replace("÷", "/"))
                    display_var.set(str(result))
                except:
                    display_var.set("错误")
            else:
                display_var.set(current + value)

        # 按钮布局
        buttons = [
            ("7", 1, 0), ("8", 1, 1), ("9", 1, 2), ("÷", 1, 3),
            ("4", 2, 0), ("5", 2, 1), ("6", 2, 2), ("×", 2, 3),
            ("1", 3, 0), ("2", 3, 1), ("3", 3, 2), ("-", 3, 3),
            ("0", 4, 0), (".", 4, 1), ("C", 4, 2), ("+", 4, 3),
            ("=", 5, 0, 4)  # 跨4列
        ]

        # 创建按钮
        for btn in buttons:
            text, row, col = btn[0], btn[1], btn[2]
            colspan = btn[3] if len(btn) > 3 else 1

            btn = ttk.Button(
                calc,
                text=text,
                command=lambda t=text: button_click(t),
                width=5,
                padding=5
            )
            btn.grid(
                row=row,
                column=col,
                columnspan=colspan,
                padx=2,
                pady=2,
                sticky="nsew"
            )

        # 设置网格权重，让按钮自适应大小
        for i in range(6):
            calc.grid_rowconfigure(i, weight=1)
        for i in range(4):
            calc.grid_columnconfigure(i, weight=1)

    def start_backup_scheduler(self):
        """启动    启动自动备份调度器（按分钟间隔执行）
        """
        self._stop_backup = threading.Event()  # 线程停止信号

        def backup_scheduler():
            while not self._stop_backup.is_set():  # 检查停止信号
                try:
                    # 检查是否启用自动备份
                    if CONFIG["backup"]["enabled"].lower() == "true":
                        # 获取配置的分钟间隔（确保有效）
                        try:
                            interval = int(CONFIG["backup"]["interval_minutes"])
                            if interval <= 0:
                                raise ValueError("间隔必须为正整数")
                        except (ValueError, KeyError) as e:
                            print(f"备份间隔配置错误：{e}，使用默认30分钟")
                            interval = 30  # 兜底默认值

                        # 执行备份（调用类内的 perform_backup 方法）
                        success = self.perform_backup()
                        if success:
                            print(f"自动备份完成，下次备份将在{interval}分钟后执行")
                        else:
                            print("自动备份失败，将在1分钟后重试")
                            interval = 1  # 备份失败时缩短重试间隔

                        # 等待指定间隔（使用事件等待，支持中途停止）
                        self._stop_backup.wait(interval * 60)
                    else:
                        # 未启用时，每10分钟检查一次配置
                        self._stop_backup.wait(600)  # 10分钟 = 600秒
                except Exception as e:
                    print(f"备份调度错误：{str(e)}")
                    self._stop_backup.wait(60)  # 出错时等待1分钟重试

        # 启动后台线程
        threading.Thread(target=backup_scheduler, daemon=True).start()

    def create_tray_icon(self):
        """创建托盘图标（保持原有逻辑）"""


        def create_default_icon():
            width, height = 64, 64
            image = Image.new('RGB', (width, height), color='blue')
            draw = ImageDraw.Draw(image)
            draw.text((10, 20), "姐妹花", fill='white')
            return image

        try:
            # 使用with语句确保文件会被自动关闭
            with Image.open(resource_path("f.ico")) as img:
                # 复制图片数据到新的Image对象，避免原文件句柄被占用
                image = img.copy()
        except:
            image = create_default_icon()

            # 定义新建销售的处理函数
        def handle_new_sale(icon=None, item=None):
            # 显示主窗口
            self.show_window(icon, item)
            # 调用SalesPage的new_sale方法
            self.sales_page.new_sale()

        menu = Menu(
            MenuItem("新建销售",handle_new_sale),
            MenuItem("显示窗口", self.show_window),
            MenuItem("退出程序", self.on_close)
        )

        return Icon("姐妹花销售系统", image, "姐妹花销售系统", menu=menu)

    def show_window(self, icon=None, item=None):
        """显示窗口并彻底清理托盘图标"""
        self.root.deiconify()
        self.root.lift()

        with self.tray_lock:
            # 强制停止所有可能存在的托盘实例
            if self.tray_icon:
                try:
                    self.tray_icon.stop()
                except:
                    pass
                self.tray_icon = None

            # 即使传递了icon参数也强制清理
            if icon:
                try:
                    icon.stop()
                except:
                    pass

            self.tray_running = False  # 确保状态重置

    # 修改最小化到托盘的方法，确保创建托盘图标
    def minimize_to_tray(self):
        """最小化到托盘"""
        self.root.withdraw()
        # 显示系统通知
        self.show_system_notification(
            "姐妹花销售系统",
            "系统已最小化到系统托盘"
        )
        with self.tray_lock:
            # 确保之前的托盘已停止再创建新的
            if self.tray_running:
                try:
                    self.tray_icon.stop()
                except:
                    pass
                self.tray_running = False

            # 启动新的托盘线程
            threading.Thread(target=self.start_tray, daemon=True).start()

    def start_tray(self):
        """启动托盘图标，优化资源释放"""
        try:
            with self.tray_lock:
                self.tray_running = True
                self.tray_icon = self.create_tray_icon()

            if self.tray_icon:
                self.tray_icon.run()  # 运行托盘主循环
        except Exception as e:
            print(f"托盘运行错误: {e}")
        finally:
            with self.tray_lock:
                self.tray_running = False
                # 强制释放图标资源
                if hasattr(self.tray_icon, '_icon') and hasattr(self.tray_icon._icon, 'handle'):
                    try:
                        windll.user32.DestroyIcon(self.tray_icon._icon.handle)
                    except:
                        pass
                self.tray_icon = None

    def cleanup_tray_resources(self):
        """强制清理托盘资源，解决句柄无效问题"""
        if self.tray_icon:
            try:
                # 停止托盘
                self.tray_icon.stop()
                # 对于Windows系统，强制释放图标句柄
                if hasattr(self.tray_icon, '_hwnd') and self.tray_icon._hwnd:
                    ctypes.windll.user32.DestroyWindow(self.tray_icon._hwnd)
            except Exception as e:
                print(f"清理托盘资源错误: {e}")
            finally:
                self.tray_icon = None

    def on_close_request(self):
        """处理关闭请求"""
        # 加载保存的偏好设置
        choice, remember = load_exit_preference()

        # 如果有记住的选项，则直接执行
        if remember and choice:
            if choice == "exit":
                self.on_close()
            elif choice == "tray":
                self.minimize_to_tray()
            return

        # 没有记住的选项，显示选择对话框
        self.show_exit_dialog()

    def show_exit_dialog(self):
        """显示退出选项对话框"""
        dialog = tk.Toplevel(self.root)
        dialog.title("退出选项")
        dialog.transient(self.root)
        dialog.geometry("300x250")
        dialog.resizable(False, False)
        dialog.grab_set()  # 模态对话框

        # 居中显示
        dialog.update_idletasks()
        width = dialog.winfo_width()
        height = dialog.winfo_height()
        x = (self.root.winfo_width() // 2) - (width // 2) + self.root.winfo_x()
        y = (self.root.winfo_height() // 2) - (height // 2) + self.root.winfo_y()
        dialog.geometry(f"+{x}+{y}")

        # 选项变量
        exit_choice = tk.StringVar(value="exit")
        remember_var = tk.BooleanVar(value=False)

        # 对话框内容
        tk.Label(dialog, text="请选择操作:").pack(pady=10)

        frame = tk.Frame(dialog)
        frame.pack(fill=tk.X, padx=20)
        tk.Radiobutton(frame, text="退出程序", variable=exit_choice, value="exit").pack(anchor=tk.W)
        tk.Radiobutton(frame, text="最小化到系统托盘", variable=exit_choice, value="tray").pack(anchor=tk.W)

        tk.Checkbutton(dialog, text="记住此选项", variable=remember_var).pack(pady=5)

        # 按钮框
        btn_frame = tk.Frame(dialog)
        btn_frame.pack(fill=tk.X, padx=20, pady=10)

        def on_confirm():
            choice = exit_choice.get()
            remember = remember_var.get()
            # 保存偏好设置
            save_exit_preference(choice, remember)

            # 关键修改：先销毁对话框，再执行后续操作
            dialog.destroy()

            # 执行选择的操作
            if choice == "exit":
                self.on_close()  # 此时对话框已销毁，再销毁主窗口
            else:
                self.minimize_to_tray()

        ttk.Button(btn_frame, text="确定", command=on_confirm).pack(fill=tk.X)

        # 允许关闭对话框（点击X时）
        def on_dialog_close():
            dialog.destroy()

        dialog.protocol("WM_DELETE_WINDOW", on_dialog_close)

    def perform_backup(self):
        """执行数据库备份，备份后只保留最新的1个文件"""
        try:
            # 确保备份目录存在
            backup_dir = CONFIG["backup"]["path"]
            os.makedirs(backup_dir, exist_ok=True)

            # 生成备份文件名（包含时间戳）
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            backup_filename = f"backup_{timestamp}.db"
            backup_path = os.path.join(backup_dir, backup_filename)

            # 复制数据库文件进行备份
            shutil.copy2(DB_PATH, backup_path)

            # 关键：备份完成后，只保留最新的1个备份（删除上一个及更早的）
            self.cleanup_old_backups(backup_dir)

            print(f"自动备份成功: {backup_path}")
            return True
        except Exception as e:
            print(f"备份失败: {str(e)}")
            return False

    def cleanup_old_backups(self, backup_dir):
        """只保留最新的1个备份文件，删除所有旧备份"""
        try:
            # 获取所有备份文件（筛选自动备份的文件）
            backup_files = [
                f for f in os.listdir(backup_dir)
                if f.startswith("backup_") and f.endswith(".db")
            ]
            if not backup_files:
                return  # 没有备份文件，无需清理

            # 按文件名中的时间戳排序（最新的备份在最后）
            # 文件名格式：backup_YYYYMMDD_HHMMSS.db，可直接按字符串排序
            backup_files.sort()

            # 只保留最后1个（最新的），删除其他所有旧备份
            if len(backup_files) > 1:
                for old_file in backup_files[:-1]:  # 取除最后一个外的所有文件
                    old_path = os.path.join(backup_dir, old_file)
                    os.remove(old_path)
                    print(f"已删除旧备份: {old_file}")
        except Exception as e:
            print(f"清理旧备份失败: {str(e)}")

    def start_test_server(self):
        """启动test_server服务器线程（改为后台线程）"""

        def run_server():
            try:
                test_server.start_server
            except Exception as e:
                print(f"服务器启动失败: {str(e)}")

        # 创建并启动后台线程，daemon=True确保主线程退出时子线程自动结束
        threading.Thread(target=run_server, daemon=True).start()

    def _get_last_push_id(self, conn, table_name):
        """获取指定表（仅 inventory）上次推送的最大 item_id"""
        if table_name != "inventory":
            return self._get_last_push_time(conn, table_name)  # 其他表仍用时间

        cur = conn.cursor()
        cur.execute(f"""
            SELECT last_push_time FROM {CONFIG['database']['last_push_table']} 
            WHERE table_name = ?
        """, (table_name,))
        result = cur.fetchone()[0]
        # 首次推送时，last_push_time 是 '2000-01-01 00:00:00'，转为 0
        return int(result) if result.isdigit() else 0

    def _update_last_push_id(self, conn, table_name):
        """更新 inventory 表的最后推送 item_id 为当前最大 item_id"""
        if table_name != "inventory":
            return self._update_last_push_time(conn, table_name)  # 其他表仍用时间

        cur = conn.cursor()
        # 查询当前 inventory 表的最大 item_id
        cur.execute("SELECT COALESCE(MAX(item_id), 0) FROM inventory")
        max_id = cur.fetchone()[0]
        # 用 last_push_time 字段存储 max_id（字符串形式）
        cur.execute(f"""
            UPDATE {CONFIG['database']['last_push_table']} 
            SET last_push_time = ? WHERE table_name = ?
        """, (str(max_id), table_name))
        conn.commit()

    def _get_last_push_time(self, conn, table_name):
        """获取指定表的最后推送时间"""
        cur = conn.cursor()
        cur.execute(f"""
            SELECT last_push_time FROM {CONFIG['database']['last_push_table']} 
            WHERE table_name = ?
        """, (table_name,))
        return cur.fetchone()[0]

    def _update_last_push_time(self, conn, table_name):
        """更新指定表的最后推送时间为当前时间"""
        cur = conn.cursor()
        now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        cur.execute(f"""
            UPDATE {CONFIG['database']['last_push_table']} 
            SET last_push_time = ? WHERE table_name = ?
        """, (now, table_name))
        conn.commit()

    # 修改MainApp类中的_push_with_retry方法，增强错误处理
    def _push_with_retry(self, payload, max_retries=3, retry_delay=1):
        """带重试机制的推送函数，优化错误处理"""
        session = requests.Session()
        # 设置连接超时和读取超时
        session.timeout = (5, 10)  # 连接超时5秒，读取超时10秒
        try:
            for attempt in range(max_retries):
                try:
                    resp = session.post(
                        self.mobile_endpoint,
                        headers={"Content-Type": "application/json"},
                        data=json.dumps(payload, ensure_ascii=False),  # 确保中文正常传输
                        timeout=session.timeout,
                        verify=False  # 关闭SSL证书验证（仅测试用）
                    )
                    resp.raise_for_status()  # 触发HTTP错误状态码的异常
                    print(f"推送成功，服务器响应: {resp.text}")
                    return True
                except requests.exceptions.RequestException as e:
                    print(f"推送尝试 {attempt + 1}/{max_retries} 失败: {str(e)}")
                    if attempt < max_retries - 1:
                        time.sleep(retry_delay * (2 ** attempt))  # 指数退避重试
            print(f"达到最大重试次数({max_retries})，推送失败")
            return False
        finally:
            session.close()  # 确保会话关闭

    # 修改推送循环中的数据准备逻辑，确保数据格式正确
    def _push_loop(self):
        """优化后的推送循环：增量推送+短期连接+重试机制"""
        while not self._stop_push.is_set():
            conn = None
            try:
                conn = sqlite3.connect(DB_PATH)
                conn.row_factory = sqlite3.Row  # 启用行工厂，方便获取字段名
                cur = conn.cursor()

                # 1. 增量读取 sales 表
                last_sales_time = self._get_last_push_time(conn, "sales")
                cur.execute("""
                    SELECT sale_id, datetime, total_due, total_paid, is_member
                    FROM sales
                    WHERE datetime > ?
                    ORDER BY datetime
                """, (last_sales_time,))
                sales = [dict(row) for row in cur.fetchall()]  # 利用row_factory直接转为字典

                # 2. 增量读取 inventory 表
                last_inv_id = self._get_last_push_id(conn, "inventory")
                cur.execute("""
                    SELECT item_id, category, name, price, member_price, remark
                    FROM inventory
                    WHERE item_id > ?  
                    ORDER BY item_id
                """, (last_inv_id,))
                inventory = [dict(row) for row in cur.fetchall()]

                # 3. 读取 accounting 数据
                last_acc_time = self._get_last_push_time(conn, "accounting")
                cur.execute("""
                    SELECT sale_id, datetime, total_paid, is_member
                    FROM sales
                    WHERE datetime > ? AND total_paid <= total_due 
                    ORDER BY datetime
                """, (last_acc_time,))
                accounting = [dict(row) for row in cur.fetchall()]

                # 仅当有数据时推送
                if sales or inventory or accounting:
                    # 1. 确定需要推送的类型
                    types_to_push = []
                    if sales:
                        types_to_push.append("sales")
                    if inventory:
                        types_to_push.append("inventory")
                    if accounting:
                        types_to_push.append("accounting")
                    # 按要求排序：sales -> inventory -> accounting
                    types_to_push.sort(key=lambda x: ["sales", "inventory", "accounting"].index(x))

                    # 2. 单独推送每个类型，每个最多重试3次（含首次）
                    all_single_success = True
                    for data_type in types_to_push:
                        # 构建仅包含当前类型的payload
                        single_payload = {
                            data_type: locals()[data_type],  # 获取对应的数据列表（sales/inventory/accounting）
                            "push_time": datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                        }
                        # 调用带重试的推送（max_retries=3确保共3次尝试）
                        if not self._push_with_retry(single_payload, max_retries=3):
                            print(f"❌ {data_type} 单独推送最终失败")
                            all_single_success = False

                    # 3. 联合推送所有类型
                    combined_payload = {
                        "sales": sales if "sales" in types_to_push else [],
                        "inventory": inventory if "inventory" in types_to_push else [],
                        "accounting": accounting if "accounting" in types_to_push else [],
                        "push_time": datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                    }
                    combined_success = self._push_with_retry(combined_payload)

                    # 4. 所有推送成功后才更新最后推送时间/ID
                    if all_single_success and combined_success:
                        if "sales" in types_to_push:
                            self._update_last_push_time(conn, "sales")
                        if "inventory" in types_to_push:
                            self._update_last_push_id(conn, "inventory")
                        if "accounting" in types_to_push:
                            self._update_last_push_time(conn, "accounting")
                        print(f"✅ 全部推送完成: 单独推送{types_to_push}，联合推送成功")
                    else:
                        print("❌ 部分推送失败，不更新最后推送状态")

            except Exception as e:
                print(f"推送循环错误: {str(e)}")
            finally:
                if conn:
                    conn.close()

            # 等待下一次推送
            self._stop_push.wait(self.push_interval)  # 使用事件等待更优雅

    def update_clock(self):
        now = datetime.now()
        wk = ["星期一", "星期二", "星期三", "星期四", "星期五", "星期六", "星期日"][now.weekday()]
        s = now.strftime("%Y年%m月%d日 %H:%M:%S")
        self.time_label.config(text=f"{wk} {s}")
        if self._running:
            self.root.after(1000, self.update_clock)

    def on_close(self, icon=None, item=None):
        # 1. 询问是否退出程序
        if not messagebox.askyesno("确认退出", "是否要退出程序？"):
            return  # 用户选择不退出，直接返回

        # 2. 执行备份操作
        backup_success = self.perform_backup()
        if not backup_success:
            messagebox.showerror("备份失败", "程序退出前备份数据库失败，请手动备份！")

        # 3. 执行原有退出逻辑
        self._running = False
        try:
            pygame.mixer.quit()
        except:
            pass
        # 关键修复2：彻底清理托盘资源
        with self.tray_lock:
            if self.tray_icon:
                self.tray_icon.stop()
                self.tray_icon = None
            self.tray_running = False
        # 停止推送和备份线程
        if hasattr(self, '_stop_push'):
            self._stop_push.set()
        if hasattr(self, '_stop_backup'):
            self._stop_backup.set()

        self.root.destroy()

class SalesPage:
    def __init__(self, parent, main_app):
        self.main_app = main_app
        self.frame = tk.Frame(parent)

        # 按钮行
        btnf = tk.Frame(self.frame)
        btnf.pack(fill=tk.X, pady=5)
        ttk.Button(btnf, text="新建销售", command=self.new_sale).pack(side=tk.LEFT, padx=5)
        ttk.Button(btnf, text="账单查询", command=self.query_sales).pack(side=tk.LEFT, padx=5)
        ttk.Button(btnf, text="退款", command=self.refund_sale).pack(side=tk.LEFT, padx=5)
        ttk.Button(btnf, text="导出销售表", command=self.export_sales).pack(side=tk.LEFT, padx=5)
        ttk.Button(btnf, text="叫号", command=self.call_number).pack(side=tk.LEFT, padx=5)
        ttk.Button(btnf, text="销量表", command=self.show_sales_top5).pack(side=tk.LEFT, padx=5)

        # 主表格
        cols = ("序号", "时间", "菜名", "数量", "应付", "实付", "会员", "备注")
        self.tree, container = make_table(self.frame, cols)
        container.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)

        self.order_count_label = ttk.Label(
            self.frame,
            text="今日订单数：0",  # 初始默认值
        )
        self.order_count_label.pack(pady=2, padx=5, anchor="w")  # 靠左侧显示，在表格下方

        self.refresh_today()

    def refresh_today(self):
        self.tree.delete(*self.tree.get_children())
        today = date.today().strftime("%Y-%m-%d")
        conn2 = sqlite3.connect(DB_PATH)
        c2 = conn2.cursor()
        c2.execute(
            "SELECT sale_id, datetime, total_due, total_paid, is_member FROM sales WHERE datetime LIKE ? ORDER BY datetime",
            (today + '%',))
        sales_records = c2.fetchall()  # 仅调用一次 fetchall()，获取所有今日记录
        today_order_count = len(sales_records)
        self.order_count_label.config(text=f"今日订单数：{today_order_count}")

        # 关键修复：遍历已获取的 sales_records，而非再次调用 c2.fetchall()
        for idx, (sid, dt, td, tp, im) in enumerate(sales_records, 1):  # 这里改了！
            c2.execute("SELECT name, quantity, remark FROM sale_items WHERE sale_id=?", (sid,))
            items = c2.fetchall()
            names = ", ".join(r[0] for r in items)
            qtys = ", ".join(str(r[1]) for r in items)
            rems = ", ".join(r[2] for r in items if r[2])
            self.tree.insert('', 'end',
                             values=(idx, dt, names, qtys, f"{td:.2f}", f"{tp:.2f}", "是" if im else "否", rems))
        conn2.close()

    def new_sale(self):
        top = tk.Toplevel(self.frame)
        top.title("新建销售")
        top.transient(self.frame)
        configure_scaling_and_font(top)

        # 1) 菜品类别
        tk.Label(top, text="菜品类别:").grid(row=0, column=0, padx=5, pady=5, sticky='e')
        conn2 = sqlite3.connect(DB_PATH)
        c2 = conn2.cursor()
        c2.execute("SELECT DISTINCT category FROM inventory")
        cats = [r[0] for r in c2.fetchall()]
        conn2.close()
        cat_var = tk.StringVar()
        cat_cb = ttk.Combobox(top, textvariable=cat_var, values=cats, state='readonly')
        cat_cb.grid(row=0, column=1)

        # 2) 菜品名称
        tk.Label(top, text="菜品名称:").grid(row=1, column=0, padx=5, pady=5, sticky='e')
        item_var = tk.StringVar()
        item_cb = ttk.Combobox(top, textvariable=item_var, state='readonly')
        item_cb.grid(row=1, column=1)

        def load_items(e=None):
            c = cat_var.get()
            conn3 = sqlite3.connect(DB_PATH)
            c3 = conn3.cursor()
            c3.execute("SELECT name FROM inventory WHERE category=?", (c,))
            item_cb['values'] = [r[0] for r in c3.fetchall()]
            conn3.close()
            item_var.set("")

        cat_cb.bind("<<ComboboxSelected>>", load_items)

        # 3) 数量
        tk.Label(top, text="数量:").grid(row=2, column=0, padx=5, pady=5, sticky='e')
        qty_e = tk.Entry(top)
        qty_e.grid(row=2, column=1)
        qty_e.insert(0, "1")

        # 4) 备注
        tk.Label(top, text="备注:").grid(row=3, column=0, padx=5, pady=5, sticky='e')
        remark_e = tk.Entry(top)
        remark_e.grid(row=3, column=1)

        # 新增：折扣输入框
        tk.Label(top, text="折扣(%):").grid(row=4, column=0, padx=5, pady=5, sticky='e')
        discount_var = tk.StringVar(value="100")  # 默认不打折
        discount_e = tk.Entry(top, textvariable=discount_var)
        discount_e.grid(row=4, column=1)

        # 5) 会员号与余额显示
        mem_flag = False
        mem_phone = ""
        mem_balance = 0.0

        tk.Label(top, text="会员号:").grid(row=5, column=0, padx=5, pady=5, sticky='e')
        phone_lbl = tk.Label(top, text="—")
        phone_lbl.grid(row=5, column=1, sticky='w')
        tk.Label(top, text="余额:").grid(row=6, column=0, padx=5, pady=5, sticky='e')
        balance_lbl = tk.Label(top, text="0.00")
        balance_lbl.grid(row=6, column=1, sticky='w')

        def select_member():
            nonlocal mem_flag, mem_phone, mem_balance
            mtop = tk.Toplevel(top)
            mtop.title("选择会员")
            configure_scaling_and_font(mtop)
            tk.Label(mtop, text="手机号:").pack(padx=5, pady=5)
            phone_e2 = tk.Entry(mtop)
            phone_e2.pack(padx=5)
            # 修改表格列，增加"加入日期"
            tree_m, cont_m = make_table(mtop, ("手机号", "加入日期", "余额", "备注"))
            cont_m.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)

            def do_search():
                tel = phone_e2.get().strip()
                conn4 = sqlite3.connect(DB_PATH)
                c4 = conn4.cursor()
                # 查询时包含join_date字段
                c4.execute("SELECT phone, join_date, balance, remark FROM members WHERE phone=?", (tel,))
                rows = c4.fetchall()
                conn4.close()
                tree_m.delete(*tree_m.get_children())
                # 显示时包含加入日期（只显示年月日）
                for ph, join_date, bal, rm in rows:
                    date_str = join_date.split()[0] if join_date else ""
                    tree_m.insert('', tk.END, values=(ph, date_str, f"{bal:.2f}", rm or ""))

            ttk.Button(mtop, text="查找", command=do_search).pack(pady=5)

            def confirm():
                nonlocal mem_flag, mem_phone, mem_balance
                sel = tree_m.selection()
                if not sel: return messagebox.showerror("错误", "请选择会员")
                vals = tree_m.item(sel[0])['values']
                mem_phone, mem_balance = vals[0], float(vals[2])  # 注意余额列索引变为2
                mem_flag = True
                phone_lbl.config(text=mem_phone)
                balance_lbl.config(text=f"{mem_balance:.2f}")

                # 刷新表格中的金额为会员价乘以折扣
                for idx, item in enumerate(items):
                    name, pr, mp, qty, dis, _, rem = item
                    discount_ratio = float(dis) / 100
                    new_amt = mp * qty * discount_ratio
                    items[idx] = (name, pr, mp, qty, dis, new_amt, rem)
                    iid = tree2.get_children()[idx]
                    tree2.item(iid, values=(
                        name, f"{pr:.2f}", f"{mp:.2f}", qty,
                        f"{dis:.1f}", f"{new_amt:.2f}", rem
                    ))

                update_totals()  # 刷新实付总和
                mtop.destroy()

            ttk.Button(mtop, text="确认", command=confirm).pack(pady=5)

        ttk.Button(top, text="选择会员", command=select_member).grid(row=7, column=0, columnspan=2, pady=5)

        # 6) 明细表格
        cols2 = ("名称", "原价", "会员价", "数量", "折扣(%)", "实付", "备注")  # 增加折扣列显示
        tree2, cont2 = make_table(top, cols2)
        cont2.grid(row=0, column=2, rowspan=12, padx=10, pady=5)

        # 存储菜品信息，格式: (名称, 原价, 会员价, 数量, 折扣, 实付, 备注)
        items = []

        # 新增：创建右键菜单
        def create_context_menu():
            context_menu = tk.Menu(top, tearoff=0)
            context_menu.add_command(label="删除", command=delete_selected_item)
            return context_menu

        # 新增：删除选中项
        def delete_selected_item():
            selected = tree2.selection()
            if not selected:
                return

            # 获取选中项的索引
            indices = [tree2.index(item) for item in selected]
            # 从后往前删除，避免索引变化影响
            for idx in sorted(indices, reverse=True):
                # 从数据列表中删除
                if 0 <= idx < len(items):
                    items.pop(idx)
                # 从表格中删除
                tree2.delete(selected[0])  # 每次删除第一个选中项

            update_totals()  # 更新总金额

        # 新增：显示右键菜单
        def show_context_menu(event):
            item = tree2.identify_row(event.y)
            if item:
                tree2.selection_set(item)
                context_menu.post(event.x_root, event.y_root)

        # 创建右键菜单并绑定事件
        context_menu = create_context_menu()
        tree2.bind("<Button-3>", show_context_menu)

        def get_discount():
            """获取折扣值，确保在0-100之间"""
            try:
                discount = float(discount_var.get() or 0)
                return max(0, min(100, discount))  # 保留百分比格式
            except:
                return 100.0  # 无效输入时默认不打折

        def update_totals():
            """更新总金额（总实付为所有菜品金额总和）"""
            total_due = sum(item[1] * item[3] for item in items)  # 应付金额（原价）
            total_paid = sum(item[5] for item in items)  # 实付金额（已包含各自折扣）
            due_lbl.config(text=f"{total_due:.2f}")
            paid_lbl.config(text=f"{total_paid:.2f}")

        def add_item():
            name = item_var.get().strip()
            qty = qty_e.get().strip()
            rem = remark_e.get().strip()
            if not name or not qty.isdigit():
                return messagebox.showerror("错误", "请输入菜品和数量")
            q = int(qty)
            conn5 = sqlite3.connect(DB_PATH)
            c5 = conn5.cursor()
            c5.execute("SELECT price, member_price FROM inventory WHERE name=?", (name,))
            row = c5.fetchone()
            conn5.close()
            if not row: return messagebox.showerror("错误", "菜品不存在")
            pr, mp = row

            # 获取当前折扣（仅对本次加入的菜品生效）
            current_discount = get_discount()
            discount_ratio = current_discount / 100

            # 计算价格：考虑会员和当前折扣
            price_to_use = mp if mem_flag else pr
            amt = price_to_use * q * discount_ratio

            # 合并同名同备注记录（保持相同折扣才合并）
            for idx, (n, p1, p2, qq, dis, _, r) in enumerate(items):
                if n == name and r == rem and dis == current_discount:
                    new_q = qq + q
                    new_amt = price_to_use * new_q * discount_ratio
                    items[idx] = (n, p1, p2, new_q, dis, new_amt, r)
                    iid = tree2.get_children()[idx]
                    tree2.item(iid, values=(n, f"{p1:.2f}", f"{p2:.2f}", new_q, f"{dis:.1f}", f"{new_amt:.2f}", r))
                    update_totals()
                    return

            # 新增菜品记录
            items.append((name, pr, mp, q, current_discount, amt, rem))
            tree2.insert('', tk.END, values=(
                name,
                f"{pr:.2f}",
                f"{mp:.2f}",
                q,
                f"{current_discount:.1f}",  # 显示折扣
                f"{amt:.2f}",
                rem
            ))
            update_totals()

        ttk.Button(top, text="加入", command=add_item).grid(row=8, column=0, columnspan=2, pady=5)

        tk.Label(top, text="总应付:").grid(row=9, column=0, sticky='e')
        due_lbl = tk.Label(top, text="0.00")
        due_lbl.grid(row=9, column=1, sticky='w')
        tk.Label(top, text="总实付:").grid(row=10, column=0, sticky='e')
        paid_lbl = tk.Label(top, text="0.00")
        paid_lbl.grid(row=10, column=1, sticky='w')

        def submit():
            nonlocal mem_flag, mem_phone, mem_balance
            if not items:
                return messagebox.showerror("错误", "请先加入菜品")

            # 会员状态恢复机制
            if mem_flag and phone_lbl.cget("text") in ("", "—"):
                mem_flag = False
                mem_phone = ""
                mem_balance = 0.0
                balance_lbl.config(text="0.00")
                update_totals()

            # 刷新金额以与状态一致
            update_totals()

            now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            td = sum(i[1] * i[3] for i in items)  # 应付金额（原价）
            tp = sum(i[5] for i in items)  # 实付金额（已包含各自折扣）
            im = 1 if mem_flag else 0

            if mem_flag and mem_balance < tp:
                return messagebox.showerror("错误", f"会员余额不足！当前余额: {mem_balance:.2f}，需支付: {tp:.2f}")

            try:
                conn6 = sqlite3.connect(DB_PATH)
                c6 = conn6.cursor()
                c6.execute("""
                            INSERT INTO sales 
                            (datetime, total_due, total_paid, is_member, member_phone) 
                            VALUES (?, ?, ?, ?, ?)
                        """, (now, td, tp, im, mem_phone if mem_flag else None))
                sid = c6.lastrowid
                for n, p1, p2, qq, _, amt, r in items:
                    c6.execute(
                        "INSERT INTO sale_items (sale_id,category,name,price,quantity,remark) VALUES (?,?,?,?,?,?)",
                        (sid, cat_var.get(), n, p1, qq, r))
                if mem_flag:
                    c6.execute("UPDATE members SET balance=? WHERE phone=?", (mem_balance - tp, mem_phone))
                conn6.commit()
                messagebox.showinfo("成功", "销售记录已成功提交！")
                top.destroy()
                self.refresh_today()  # 刷新今日销售列表
                self.main_app.members_page.refresh()   # 添加这一行
                self.main_app.update_sales_progress()
                # 记录销售次数（使用数据库存储）
                self.record_sale_streak()
            except Exception as e:
                messagebox.showerror("错误", f"提交失败: {str(e)}")

        ttk.Button(top, text="提交", command=submit).grid(row=11, column=0, columnspan=2, pady=5)

    def record_sale_streak(self):
        conn = sqlite3.connect(DB_PATH)
        cur = conn.cursor()
        cur.execute(
            "CREATE TABLE IF NOT EXISTS user_achievements (id INTEGER PRIMARY KEY, sale_count INTEGER DEFAULT 0)")
        cur.execute("INSERT OR IGNORE INTO user_achievements (id, sale_count) VALUES (1, 0)")
        cur.execute("UPDATE user_achievements SET sale_count = sale_count + 1 WHERE id = 1")
        cur.execute("SELECT sale_count FROM user_achievements WHERE id = 1")
        count = cur.fetchone()[0]
        conn.commit()
        conn.close()

        # 成就触发
        achievements = {
            1:"新手",
            10: "初露锋芒",
            50: "销售能手",
            100: "销售冠军",
            500: "金牌店长"
        }
        if count in achievements:
            messagebox.showinfo("解锁成就", f"恭喜您获得「{achievements[count]}」称号！已完成{count}笔销售～")

    def query_sales(self):
        top = tk.Toplevel(self.frame)
        top.title("账单查询")
        top.transient(self.frame)
        configure_scaling_and_font(top)

        tk.Label(top, text="日期 (YYYY-MM-DD):").pack(pady=5)
        date_e = tk.Entry(top)
        date_e.pack(pady=5)

        cols = ("序号", "时间", "菜名", "数量", "应付", "实付", "会员", "备注")
        tree_q, cont_q = make_table(top, cols)
        cont_q.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)

        sum_lbl = tk.Label(top, text="实付总计: 0.00")
        sum_lbl.pack(pady=5)

        def do_query():
            d = date_e.get().strip()
            if not d: return messagebox.showerror("错误", "请输入日期")
            tree_q.delete(*tree_q.get_children())
            conn2 = sqlite3.connect(DB_PATH)
            c2 = conn2.cursor()
            c2.execute(
                "SELECT sale_id, datetime, total_due, total_paid, is_member FROM sales WHERE datetime LIKE ? ORDER BY datetime",
                (d + '%',))
            total = 0.0
            for idx, (sid, dt, td, tp, im) in enumerate(c2.fetchall(), 1):
                c2.execute("SELECT name, quantity, remark FROM sale_items WHERE sale_id=?", (sid,))
                items = c2.fetchall()
                names = ", ".join(r[0] for r in items)
                qtys = ", ".join(str(r[1]) for r in items)
                rems = ", ".join(r[2] for r in items if r[2])
                tree_q.insert('', 'end',
                              values=(idx, dt, names, qtys, f"{td:.2f}", f"{tp:.2f}", "是" if im else "否", rems))
                total += tp
            conn2.close()
            sum_lbl.config(text=f"实付总计: {total:.2f}")
        # 创建按钮容器框架，用于横向排列按钮
        btn_frame = tk.Frame(top)
        btn_frame.pack(pady=5)

            # 将按钮放入框架，使用side=tk.LEFT实现横向排列
        ttk.Button(btn_frame, text="查询", command=do_query).pack(side=tk.LEFT, padx=10)
        ttk.Button(btn_frame, text="关闭", command=top.destroy).pack(side=tk.LEFT, padx=10)
        # 添加回车键绑定，按下回车时执行查询
        date_e.bind("<Return>", lambda event: do_query())

    def export_sales(self):
        fn = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel 文件", "*.xlsx")])
        if not fn: return
        rows = [self.tree.item(i)['values'] for i in self.tree.get_children()]
        wb = Workbook()
        ws = wb.active
        ws.append(["序号", "时间", "菜名", "数量", "应付", "实付", "会员", "备注"])
        for r in rows: ws.append(r)
        wb.save(fn)
        messagebox.showinfo("导出", "导出成功")

    def call_number(self):
        top = tk.Toplevel(self.frame)
        top.title("叫号")
        top.transient(self.frame)
        configure_scaling_and_font(top)

        tk.Label(top, text="序号:").grid(row=0, column=0, padx=5, pady=5)
        num_e = tk.Entry(top)
        num_e.grid(row=0, column=1, pady=5)

        tk.Label(top, text="次数:").grid(row=1, column=0, padx=5, pady=5)
        rep_e = tk.Entry(top)
        rep_e.grid(row=1, column=1, pady=5)
        rep_e.insert(0, "1")

        # 为两个输入框添加回车键绑定，按下回车时执行叫号
        num_e.bind("<Return>", lambda event: do_call())
        rep_e.bind("<Return>", lambda event: do_call())

        def speak_text(text):
            """使用pyttsx3进行文本朗读"""
            engine = pyttsx3.init()
            # 可以调整语速，范围一般是0-200
            engine.setProperty('rate', 150)
            engine.say(text)
            engine.runAndWait()

        def do_call():
            idx = num_e.get().strip()
            rpt = rep_e.get().strip()
            if not idx.isdigit() or not rpt.isdigit():
                return messagebox.showerror("错误", "请输入数字")

            # 构建要朗读的文本
            text = f"请{idx}号顾客前来取餐"
            repeat = int(rpt)

            # 启动线程进行朗读，避免界面卡顿
            def speak_task():
                for _ in range(repeat):
                    speak_text(text)
                    # 每次朗读后等待0.5秒
                    time.sleep(0.5)

            threading.Thread(target=speak_task, daemon=True).start()
            top.destroy()

        ttk.Button(top, text="确定", command=do_call).grid(row=2, columnspan=2, pady=10)


    def refund_sale(self):
        top = tk.Toplevel(self.frame)
        top.title("销售退款")
        top.transient(self.frame)
        configure_scaling_and_font(top)
        top.geometry("1200x600")

        # 日期输入
        tk.Label(top, text="日期 (YYYY-MM-DD):").pack(pady=5)
        date_e = tk.Entry(top)
        date_e.pack(pady=5)

        # 表格（新增会员手机号列，用于内部存储）
        cols = ("序号", "时间", "菜名", "数量", "应付", "实付", "会员", "会员手机号", "备注")  # 新增会员手机号列
        tree_r, cont_r = make_table(top, cols)
        cont_r.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)

        sum_lbl = tk.Label(top, text="总计: 0.00")
        sum_lbl.pack(pady=5)

        # 按钮区域
        btn_frame = tk.Frame(top)
        btn_frame.pack(pady=10)

        def do_query():
            d = date_e.get().strip()
            if not d:
                return messagebox.showerror("错误", "请输入日期")
            tree_r.delete(*tree_r.get_children())
            with sqlite3.connect(DB_PATH) as conn:
                c = conn.cursor()
                try:
                    # 查询时包含 member_phone（确保从 sales 表能获取到）
                    c.execute("""
                        SELECT sale_id, datetime, total_due, total_paid, is_member, member_phone 
                        FROM sales 
                        WHERE datetime LIKE ? 
                        ORDER BY datetime
                    """, (d + '%',))
                    total = 0.0
                    for idx, (sid, dt, td, tp, im, mem_phone) in enumerate(c.fetchall(), 1):
                        c.execute("SELECT name, quantity, remark FROM sale_items WHERE sale_id=?", (sid,))
                        items = c.fetchall()
                        names = ", ".join(r[0] for r in items)
                        qtys = ", ".join(str(r[1]) for r in items)  # 确保数量是字符串
                        rems = ", ".join(r[2] for r in items if r[2])

                        # 插入表格时，明确会员手机号为字符串（即使为空也保持字符串类型）
                        tree_r.insert('', 'end', iid=sid, values=(
                            idx,  # 序号（索引1）
                            dt,  # 时间（索引2）
                            names,  # 菜名（索引3）
                            qtys,  # 数量（索引4，字符串类型）
                            f"{td:.2f}",  # 应付（索引5）
                            f"{tp:.2f}",  # 实付（索引6）
                            "是" if im else "否",  # 会员（索引7）
                            str(mem_phone) if mem_phone else "",  # 会员手机号（索引8，强制转为字符串）
                            rems  # 备注（索引9）
                        ))
                        total += tp
                    sum_lbl.config(text=f"总计: {total:.2f}")
                except Exception as e:
                    messagebox.showerror("错误", f"查询失败: {str(e)}")

        def process_refund():
            sel = tree_r.selection()
            if not sel:
                return messagebox.showerror("错误", "请选择要退款的记录")

            vals = tree_r.item(sel[0])['values']
            # 修正索引：根据上面的插入逻辑，会员手机号在索引8
            sale_id = sel[0]  # sel[0]是选中项的iid，即数据库中的sale_id
            total_paid = float(vals[5])  # 实付金额（索引6）
            is_member = vals[6] == "是"  # 会员状态（索引7）
            # 强制转为字符串后再strip()，避免int类型错误
            member_phone = str(vals[7]).strip()  # 会员手机号（索引8）

            # 确认退款
            if not messagebox.askyesno("确认", f"确定要退还这笔交易吗？金额: {total_paid:.2f}元"):
                return

            with sqlite3.connect(DB_PATH) as conn:
                c = conn.cursor()
                try:
                    # 如果是会员，需要退还金额到会员账户
                    if is_member:
                        if not member_phone:
                            raise Exception("未找到该销售记录关联的会员信息")

                        # 验证会员是否存在
                        c.execute("SELECT phone FROM members WHERE phone=?", (member_phone,))
                        if not c.fetchone():
                            raise Exception(f"会员 {member_phone} 不存在")

                        # 更新会员余额（加上退款金额）
                        c.execute("""
                            UPDATE members 
                            SET balance = balance + ? 
                            WHERE phone = ?
                        """, (total_paid, member_phone))
                        messagebox.showinfo("成功", f"已向会员 {member_phone} 退还 {total_paid:.2f} 元到余额")
                    else:
                        messagebox.showinfo("提示", f"非会员退款，金额 {total_paid:.2f} 元，请线下处理")

                    # 删除销售记录和明细
                    c.execute("DELETE FROM sale_items WHERE sale_id = ?", (sale_id,))
                    c.execute("DELETE FROM sales WHERE sale_id = ?", (sale_id,))
                    conn.commit()

                    # 事件推送逻辑（增强顺序性和明确性）
                    from test_server import data_queue
                    # 1. 推送删除销售记录事件
                    data_queue.put(("delete_sale", [{"sale_id": sale_id}]))
                    print(f"📤 已推送销售删除事件: {sale_id}")
                    time.sleep(0.05)  # 确保事件顺序

                    # 2. 推送删除记账记录事件
                    data_queue.put(("delete_accounting", [{"sale_id": sale_id}]))
                    print(f"📤 已推送记账删除事件: {sale_id}")
                    time.sleep(0.05)

                    # 3. 推送全量刷新事件（明确携带当前时间戳，确保网页端识别为新事件）
                    from datetime import datetime
                    combined_payload = {
                        "sales": [],
                        "inventory": [],
                        "accounting": [],
                        "push_time": datetime.now().strftime("%Y-%m-%d %H:%M:%S.%f"),  # 精确到微秒，避免缓存
                        "full_refresh": True
                    }
                    data_queue.put(("combined", combined_payload))
                    print(f"📤 已推送联合刷新事件（带精确时间戳）")

                    # 关键调整：事件推送后强制刷新本地页面，并延迟极短时间确保网页端收到事件
                    messagebox.showinfo("成功", "退款操作已完成，销售记录已删除")
                    do_query()  # 刷新退款页面表格
                    time.sleep(0.1)  # 给网页端处理事件的时间
                    self.refresh_today()  # 刷新主销售页面
                    self.main_app.members_page.refresh()  # 刷新会员页面
                    self.main_app.update_sales_progress()  # 更新销售进度条
                    top.destroy()
                except Exception as e:
                    conn.rollback()
                    messagebox.showerror("错误", f"退款失败: {str(e)}")

        ttk.Button(btn_frame, text="查询", command=do_query).pack(side=tk.LEFT, padx=10)
        ttk.Button(btn_frame, text="退款", command=process_refund).pack(side=tk.LEFT, padx=10)
        ttk.Button(btn_frame, text="关闭", command=top.destroy).pack(side=tk.LEFT, padx=10)
        date_e.bind("<Return>", lambda event: do_query())

    def show_sales_top5(self):
        """显示销量Top5商品柱状图窗口"""
        top = tk.Toplevel(self.frame)
        top.title("销量Top5")
        top.transient(self.frame)
        configure_scaling_and_font(top)

        # 时间范围选择
        tk.Label(top, text="统计周期:").grid(row=0, column=0, padx=10, pady=20, sticky='e')
        period_var = tk.StringVar(value="周")
        # 给下拉框添加变量名，便于绑定事件
        period_cb = ttk.Combobox(
            top,
            textvariable=period_var,
            values=["周", "月", "年"],
            state='readonly'
        )
        period_cb.grid(row=0, column=1, padx=10, pady=20)
        # 绑定回车键到下拉框，按下回车时生成图表
        period_cb.bind("<Return>", lambda event: generate_chart())

        # 生成图表按钮
        def generate_chart():
            period = period_var.get()
            self.plot_sales_top5(period)
            top.destroy()

        ttk.Button(
            top,
            text="生成图表",
            command=generate_chart
        ).grid(row=1, column=0, columnspan=2, pady=30)

    def plot_sales_top5(self, period):
        """绘制销量Top5柱状图"""
        plt.rcParams['font.sans-serif'] = ['SimHei']
        plt.rcParams['axes.unicode_minus'] = False

        # 计算时间范围
        now = datetime.now()
        if period == "周":
            start_date = (now - timedelta(days=7)).strftime("%Y-%m-%d")
            title = f"近7天销量Top5"
        elif period == "月":
            start_date = now.strftime("%Y-%m-01")
            title = f"{now.strftime('%Y年%m月')}销量Top5"
        else:  # 年
            start_date = now.strftime("%Y-01-01")
            title = f"{now.strftime('%Y年')}销量Top5"

        # 获取当前库存中的所有商品名称
        conn = sqlite3.connect(DB_PATH)
        cursor = conn.cursor()
        cursor.execute("SELECT name FROM inventory")
        inventory_items = {row[0] for row in cursor.fetchall()}  # 转换为集合便于快速查找
        conn.close()

        # 查询销量数据
        conn = sqlite3.connect(DB_PATH)
        cursor = conn.cursor()
        cursor.execute("""
            SELECT i.name, SUM(i.quantity) as total_qty
            FROM sale_items i
            JOIN sales s ON i.sale_id = s.sale_id
            WHERE s.datetime >= ?
            GROUP BY i.name
            ORDER BY total_qty DESC
            LIMIT 5
        """, (start_date,))

        data = cursor.fetchall()
        conn.close()

        # 过滤掉不在库存中的商品
        data = [item for item in data if item[0] in inventory_items]

        if not data:
            messagebox.showinfo("提示", "没有找到销售数据或所有商品均已下架")
            return

        # 处理数据
        names = [item[0] for item in data]
        quantities = [item[1] for item in data]
        ranks = [f"Top{i + 1}" for i in range(len(data))]

        # 绘制柱状图
        plt.figure(figsize=(10, 6))
        bars = plt.bar(ranks, quantities, color='skyblue')

        # 在柱子上方标注商品名称
        for bar, name in zip(bars, names):
            height = bar.get_height()
            plt.text(
                bar.get_x() + bar.get_width() / 2.,
                height + 0.1,
                name,
                ha='center',
                va='bottom',
                rotation=0,
                fontsize=10
            )

        plt.title(title)
        plt.xlabel("排名")
        plt.ylabel("销售数量(个)")
        plt.ylim(0, max(quantities) * 1.1)  # 留出显示名称的空间
        plt.tight_layout()
        plt.show()

class MemberPage:
    def __init__(self, parent):
        self.frame = tk.Frame(parent)
        configure_scaling_and_font(self.frame)

        # 按钮和搜索
        topf = tk.Frame(self.frame)
        topf.pack(fill=tk.X, pady=5)
        ttk.Label(topf, text="搜索手机号/备注:").pack(side=tk.LEFT, padx=5)
        self.search_e = tk.Entry(topf)
        self.search_e.pack(side=tk.LEFT, padx=5)
        ttk.Button(topf, text="搜索", command=self.search).pack(side=tk.LEFT, padx=5)
        ttk.Button(topf, text="显示全部", command=self.refresh).pack(side=tk.LEFT, padx=5)
        ttk.Button(topf, text="新增会员", command=self.add_member).pack(side=tk.LEFT, padx=5)
        ttk.Button(topf, text="修改会员", command=self.edit_member).pack(side=tk.LEFT, padx=5)
        ttk.Button(topf, text="充值余额", command=self.charge_member).pack(side=tk.LEFT, padx=5)
        ttk.Button(topf, text="删除会员", command=self.delete_member).pack(side=tk.LEFT, padx=5)
        ttk.Button(topf, text="导出会员表", command=self.export_members).pack(side=tk.LEFT, padx=5)

        # 表格 - 添加"加入日期"列
        cols = ("手机号", "加入日期", "余额", "备注")  # 新增"加入日期"列
        self.tree, container = make_table(self.frame, cols)
        container.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)

        # 会员总数
        self.count_lbl = ttk.Label(self.frame, text="共 0 个会员")
        self.count_lbl.pack(pady=5)

        self.refresh()

    def refresh(self):
        self.tree.delete(*self.tree.get_children())
        conn2 = sqlite3.connect(DB_PATH)
        c2 = conn2.cursor()
        # 查询时添加join_date字段
        c2.execute("SELECT phone, join_date, balance, remark FROM members ORDER BY phone")
        rows = c2.fetchall()
        conn2.close()
        # 显示时包含加入日期
        for phone, join_date, bal, rem in rows:
            # 格式化日期显示（只显示年月日）
            date_str = join_date.split()[0] if join_date else ""
            self.tree.insert('', tk.END, values=(phone, date_str, f"{bal:.2f}", rem or ""))
        self.count_lbl.config(text=f"共 {len(rows)} 个会员")

    def search(self):
        key = self.search_e.get().strip()
        conn2 = sqlite3.connect(DB_PATH)
        c2 = conn2.cursor()
        if key:
            # 搜索查询也需要包含join_date
            c2.execute("SELECT phone, join_date, balance, remark FROM members WHERE phone LIKE ? OR remark LIKE ?",
                       (f"%{key}%", f"%{key}%"))
        else:
            c2.execute("SELECT phone, join_date, balance, remark FROM members")
        rows = c2.fetchall()
        conn2.close()
        self.tree.delete(*self.tree.get_children())
        # 显示搜索结果时包含加入日期
        for phone, join_date, bal, rem in rows:
            date_str = join_date.split()[0] if join_date else ""
            self.tree.insert('', tk.END, values=(phone, date_str, f"{bal:.2f}", rem or ""))
        self.count_lbl.config(text=f"共 {len(rows)} 个会员")

    def add_member(self):
        top = tk.Toplevel(self.frame)
        top.title("新增会员")
        top.transient(self.frame)
        configure_scaling_and_font(top)
        tk.Label(top, text="手机号:").grid(row=0, column=0, padx=5, pady=5, sticky='e')
        phone_e = tk.Entry(top)
        phone_e.grid(row=0, column=1)
        # 手机号输入框绑定回车
        phone_e.bind("<Return>", lambda event: save())
        tk.Label(top, text="初始余额:").grid(row=1, column=0, padx=5, pady=5, sticky='e')
        bal_e = tk.Entry(top)
        bal_e.grid(row=1, column=1)
        # 初始余额输入框绑定回车
        bal_e.bind("<Return>", lambda event: save())
        tk.Label(top, text="备注:").grid(row=2, column=0, padx=5, pady=5, sticky='e')
        rem_e = tk.Entry(top)
        rem_e.grid(row=2, column=1)
        # 备注输入框绑定回车
        rem_e.bind("<Return>", lambda event: save())

        def save():
            tel = phone_e.get().strip()
            try:
                bal = float(bal_e.get())
            except:
                return messagebox.showerror("错误", "余额格式错误")
            rem = rem_e.get().strip()
            if not tel: return messagebox.showerror("错误", "手机号不能为空")

            # 获取当前日期时间作为加入日期
            from datetime import datetime
            current_date = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

            conn2 = sqlite3.connect(DB_PATH)
            c2 = conn2.cursor()
            try:
                # 插入时包含join_date字段
                c2.execute("INSERT INTO members (phone, balance, remark, join_date) VALUES (?,?,?,?)",
                           (tel, bal, rem, current_date))
                conn2.commit()
                messagebox.showinfo("成功", f"会员 {tel} 添加成功")
            except sqlite3.IntegrityError:
                conn2.close()
                return messagebox.showerror("错误", "手机号已存在")
            except Exception as e:
                conn2.close()
                return messagebox.showerror("错误", f"添加失败: {str(e)}")
            finally:
                conn2.close()
            self.refresh()
            top.destroy()

        ttk.Button(top, text="保存", command=save).grid(row=3, columnspan=2, pady=10)

    def edit_member(self):
        sel = self.tree.selection()
        if not sel: return messagebox.showerror("错误", "请选择会员")
        phone_old = self.tree.item(sel[0])['values'][0]
        top = tk.Toplevel(self.frame)
        top.title("修改会员")
        top.transient(self.frame)
        configure_scaling_and_font(top)
        tk.Label(top, text="手机号:").grid(row=0, column=0, padx=5, pady=5, sticky='e')
        phone_e = tk.Entry(top)
        phone_e.grid(row=0, column=1)
        # 手机号输入框绑定回车，触发保存
        phone_e.bind("<Return>", lambda event: save())
        phone_e.insert(0, phone_old)
        tk.Label(top, text="余额:").grid(row=1, column=0, padx=5, pady=5, sticky='e')
        bal_e = tk.Entry(top)
        bal_e.grid(row=1, column=1)
        bal_e.insert(0, self.tree.item(sel[0])['values'][1])
        # 余额输入框绑定回车，触发保存
        bal_e.bind("<Return>", lambda event: save())
        tk.Label(top, text="备注:").grid(row=2, column=0, padx=5, pady=5, sticky='e')
        rem_e = tk.Entry(top)
        rem_e.grid(row=2, column=1)
        # 备注输入框绑定回车，触发保存
        rem_e.bind("<Return>", lambda event: save())
        rem_e.insert(0, self.tree.item(sel[0])['values'][2])

        def save():
            tel = phone_e.get().strip()
            try:
                bal = float(bal_e.get())
            except:
                return messagebox.showerror("错误", "余额格式错误")
            rem = rem_e.get().strip()
            conn2 = sqlite3.connect(DB_PATH)
            c2 = conn2.cursor()
            try:
                c2.execute("UPDATE members SET phone=?, balance=?, remark=? WHERE phone=?",
                           (tel, bal, rem, phone_old))
                conn2.commit()
                messagebox.showinfo("成功", f"会员 {tel} 修改成功")
            except sqlite3.IntegrityError:
                conn2.close()
                return messagebox.showerror("错误", "手机号已存在")
            except Exception as e:
                conn2.close()
                return messagebox.showerror("错误", f"修改失败: {str(e)}")
            finally:
                conn2.close()
            self.refresh()
            top.destroy()

        ttk.Button(top, text="保存", command=save).grid(row=3, columnspan=2, pady=10)

    def charge_member(self):
        sel = self.tree.selection()
        if not sel: return messagebox.showerror("错误", "请选择会员")
        phone = self.tree.item(sel[0])['values'][0]
        top = tk.Toplevel(self.frame)
        top.title("充值")
        top.transient(self.frame)
        configure_scaling_and_font(top)
        tk.Label(top, text="充值金额:").grid(row=0, column=0, padx=5, pady=5, sticky='e')
        amt_e = tk.Entry(top)
        amt_e.grid(row=0, column=1)
        # 给充值金额输入框绑定回车键，触发保存操作
        amt_e.bind("<Return>", lambda event: save())

        def save():
            try:
                amt = float(amt_e.get())
            except:
                return messagebox.showerror("错误", "金额格式错误")
            if amt <= 0: return messagebox.showerror("错误", "充值金额必须大于0")
            conn2 = sqlite3.connect(DB_PATH)
            c2 = conn2.cursor()
            try:
                c2.execute("UPDATE members SET balance=balance+? WHERE phone=?", (amt, phone))
                conn2.commit()
                messagebox.showinfo("成功", f"会员 {phone} 充值 {amt} 元成功")
            except Exception as e:
                conn2.close()
                return messagebox.showerror("错误", f"充值失败: {str(e)}")
            finally:
                conn2.close()
            self.refresh()
            top.destroy()

        ttk.Button(top, text="保存", command=save).grid(row=1, columnspan=2, pady=10)

    def delete_member(self):
        sel = self.tree.selection()
        if not sel: return messagebox.showerror("错误", "请选择会员")
        phone = self.tree.item(sel[0])['values'][0]
        if not messagebox.askyesno("确认", f"删除会员 {phone}?"): return
        conn2 = sqlite3.connect(DB_PATH)
        c2 = conn2.cursor()
        try:
            c2.execute("DELETE FROM members WHERE phone=?", (phone,))
            conn2.commit()
            messagebox.showinfo("成功", f"会员 {phone} 已删除")
        except Exception as e:
            conn2.close()
            return messagebox.showerror("错误", f"删除失败: {str(e)}")
        finally:
            conn2.close()
        self.refresh()

    def export_members(self):
        fn = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel 文件", "*.xlsx")])
        if not fn: return
        rows = [self.tree.item(i)['values'] for i in self.tree.get_children()]
        wb = Workbook()
        ws = wb.active
        # 导出时包含加入日期列
        ws.append(["手机号", "加入日期", "余额", "备注"])
        for r in rows: ws.append(r)
        wb.save(fn)
        messagebox.showinfo("导出", "已保存到：" + fn)

class AccountPage:
    def __init__(self, parent, main_app):
        self.main_app = main_app
        self.frame = tk.Frame(parent)
        configure_scaling_and_font(self.frame)

        # 按钮行
        topf = tk.Frame(self.frame)
        topf.pack(fill=tk.X, pady=5)
        ttk.Label(topf, text="日期 :").pack(side=tk.LEFT, padx=5)
        self.date_e = tk.Entry(topf)
        self.date_e.pack(side=tk.LEFT, padx=5)
        ttk.Button(topf, text="查询", command=self.search).pack(side=tk.LEFT, padx=5)
        ttk.Button(topf, text="导出账单表", command=self.export_account).pack(side=tk.LEFT, padx=5)
        ttk.Button(topf, text="删除全部账单", command=self.delete_all_bills).pack(side=tk.LEFT, padx=5)
        ttk.Button(topf, text="统计图", command=self.open_chart_window).pack(side=tk.LEFT, padx=5)

        # 表格
        cols = ("序号", "时间", "菜名", "数量", "应付", "实付", "会员", "备注")
        self.tree, container = make_table(self.frame, cols)
        container.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)

        # 总计
        bottomf = tk.Frame(self.frame)
        bottomf.pack(fill=tk.X, pady=5)
        ttk.Label(bottomf, text="总计:").pack(side=tk.LEFT, padx=5)
        self.sum_lbl = ttk.Label(bottomf, text="0.00")
        self.sum_lbl.pack(side=tk.LEFT)

        # 订单数
        ttk.Label(bottomf, text="订单数:").pack(side=tk.LEFT, padx=5)
        self.order_count_lbl = ttk.Label(bottomf, text="0")  # 初始默认值为0
        self.order_count_lbl.pack(side=tk.LEFT)

    def search(self):
        d = self.date_e.get().strip()
        if not d:
            return messagebox.showerror("错误", "请输入日期")
        if len(d) not in (4, 7, 10):
            return messagebox.showerror("错误", "日期格式错误")
        pattern = d + "%"

        # 清空表格
        self.tree.delete(*self.tree.get_children())
        total = 0.0
        today_order_count = 0

        try:
            conn2 = sqlite3.connect(DB_PATH, timeout=30, check_same_thread=False)
            c2 = conn2.cursor()
            c2.execute("SELECT sale_id, datetime, total_due, total_paid, is_member "
                       "FROM sales WHERE datetime LIKE ? ORDER BY datetime", (pattern,))
            rows = c2.fetchall()
            today_order_count = len(rows)
            for idx, (sid, dt, td, tp, im) in enumerate(rows, 1):
                c2.execute("SELECT name, quantity, remark FROM sale_items WHERE sale_id=?", (sid,))
                items = c2.fetchall()
                names = ", ".join(r[0] for r in items)
                qtys = ", ".join(str(r[1]) for r in items)
                rems = ", ".join(r[2] for r in items if r[2])
                self.tree.insert('', tk.END, values=(
                    idx, dt, names, qtys, f"{td:.2f}", f"{tp:.2f}", "是" if im else "否", rems
                ))
                total += tp
        except Exception as e:
            messagebox.showerror("数据库错误", f"查询失败：{e}")
        finally:
            conn2.close()

        self.sum_lbl.config(text=f"{total:.2f}")
        self.order_count_lbl.config(text=str(today_order_count))
        return None

    def export_account(self):
        fn = filedialog.asksaveasfilename(defaultextension=".xlsx",
                                          filetypes=[("Excel 文件", "*.xlsx")])
        if not fn:
            return
        rows = [self.tree.item(i)['values'] for i in self.tree.get_children()]
        try:
            wb = Workbook()
            ws = wb.active
            ws.append(["序号", "时间", "菜名", "数量", "应付", "实付", "会员", "备注"])
            for r in rows:
                ws.append(r)
            wb.save(fn)
            # 修复：使用messagebox而不是不存在的show_message方法
            messagebox.showinfo("导出成功", f"已保存到：{fn}")
        except Exception as e:
            messagebox.showerror("导出失败", f"文件保存失败：{e}")

    def delete_all_bills(self):
        if not messagebox.askyesno("确认 (1/3)", "删除所有账单？"): return
        if not messagebox.askyesno("确认 (2/3)", "再次确认删除？"): return
        if not messagebox.askyesno("确认 (3/3)", "最后确认：无法恢复！"): return
        try:
            conn2 = sqlite3.connect(DB_PATH, timeout=30, check_same_thread=False)
            c2 = conn2.cursor()
            c2.execute("DELETE FROM sale_items")
            c2.execute("DELETE FROM sales")
            conn2.commit()
            conn2.close()
            self.tree.delete(*self.tree.get_children())
            self.sum_lbl.config(text="0.00")
            self.order_count_lbl.config(text="0")
            self.main_app.sales_page.refresh_today()
            messagebox.showinfo("完成", "所有账单已删除")
        except Exception as e:
            messagebox.showerror("错误", f"删除失败：{e}")

    def open_chart_window(self):
        win = tk.Toplevel(self.frame)
        win.transient(self.frame)  # 设置为父窗口的临时窗口
        win.title("统计图选项")
        tk.Label(win, text="图表类型：").grid(row=0, column=0, padx=5, pady=5, sticky='e')
        chart_type = tk.StringVar(value="条形统计图")
        ttk.Combobox(win, textvariable=chart_type, values=["条形统计图", "折线统计图", "扇形统计图"],
                     state='readonly').grid(row=0, column=1)

        tk.Label(win, text="图表模式：").grid(row=1, column=0, padx=5, pady=5, sticky='e')
        chart_mode = tk.StringVar(value="单式")
        ttk.Combobox(win, textvariable=chart_mode, values=["单式", "复合"], state='readonly').grid(row=1, column=1)

        ttk.Button(win, text="生成图表",
                   command=lambda: (win.destroy(), self.plot_chart(chart_type.get(), chart_mode.get()))).grid(
            row=2, column=0, columnspan=2, pady=10)

    def plot_chart(self, chart_type, chart_mode):
        plt.rcParams['font.sans-serif'] = ['SimHei']
        plt.rcParams['axes.unicode_minus'] = False

        def get_data(d):
            # 原 get_data 函数完全保留，无任何修改
            if len(d) == 10:
                group, label = "substr(datetime,12,2)", "小时"
            elif len(d) == 7:
                group, label = "substr(datetime,1,10)", "日期"
            elif len(d) == 4:
                group, label = "substr(datetime,1,7)", "月份"
            else:
                return [], [], ""
            conn2 = sqlite3.connect(DB_PATH)
            cursor = conn2.cursor()
            cursor.execute(f"""
                SELECT {group}, SUM(total_paid)
                FROM sales
                WHERE datetime LIKE ?
                GROUP BY {group}
                ORDER BY {group}
            """, (d + '%',))
            data = cursor.fetchall()
            conn2.close()
            return [k for k, _ in data], [v for _, v in data], label

        if chart_mode == "单式":
            d = self.date_e.get().strip()
            if not d:
                return messagebox.showerror("错误", "请输入日期")
            x, y, xlabel = get_data(d)
            if not x:
                return messagebox.showinfo("提示", "无数据")
            plt.figure()
            if chart_type == "条形统计图":
                bars = plt.bar(x, y)  # 新增：用变量接收bar对象，便于后续标注
                # 新增：为每个柱子添加金额标注（保留2位小数）
                for bar in bars:
                    height = bar.get_height()  # 获取柱子高度（即金额）
                    plt.text(bar.get_x() + bar.get_width()/2.,  # 水平居中
                             height + 0.5,  # 标注在柱子顶部上方0.5单位处（避免重叠）
                             f'{height:.2f}',  # 金额格式（保留2位小数）
                             ha='center', va='bottom')  # 水平/垂直对齐方式
            elif chart_type == "折线统计图":
                # 新增：用变量接收plot对象，获取折线的(x,y)坐标
                line = plt.plot(x, y, marker='o')[0]
                # 新增：为每个点添加金额标注（保留2位小数）
                for i, (xi, yi) in enumerate(zip(x, y)):
                    plt.text(xi, yi + 5,  # 标注在点的上方1单位处
                             f'{yi:.2f}',  # 金额格式
                             ha='center', va='bottom')  # 对齐方式
            elif chart_type == "扇形统计图":
                # 新增：用变量接收pie对象，获取每个扇区的数值
                wedges, texts, autotexts = plt.pie(y, labels=x, autopct='%1.1f%%')
                # 新增：在百分比旁添加具体金额（格式：百分比\n金额）
                for autotext, yi in zip(autotexts, y):
                    original_text = autotext.get_text()  # 获取原百分比文本（如"25.0%"）
                    autotext.set_text(f'{original_text}\n{yi:.2f}')  # 追加金额
            plt.title(f"{d} 实付统计")
            if chart_type != "扇形统计图":
                plt.xlabel(xlabel)
                plt.ylabel("实付总和")
            plt.tight_layout()
            plt.show()

        elif chart_mode == "复合":
            def ask_string(title, prompt):
                # 原 ask_string 函数完全保留，无任何修改
                dialog = tk.Toplevel(self.frame)
                dialog.transient(self.frame)
                dialog.title(title)
                ttk.Label(dialog, text=prompt).pack(pady=10)
                entry = ttk.Entry(dialog)
                entry.pack(pady=5)

                result = tk.StringVar(value="")

                def on_confirm():
                    result.set(entry.get())
                    dialog.destroy()

                def on_cancel():
                    result.set("")
                    dialog.destroy()

                confirm_button = ttk.Button(dialog, text="确认", command=on_confirm)
                confirm_button.pack(side=tk.LEFT, padx=10, pady=10)

                cancel_button = ttk.Button(dialog, text="取消", command=on_cancel)
                cancel_button.pack(side=tk.RIGHT, padx=10, pady=10)

                dialog.wait_window()
                return result.get()

            d1 = ask_string("日期一", "请输入第一个日期")
            d2 = ask_string("日期二", "请输入第二个日期")
            if not d1 or not d2:
                return
            x1, y1, xlabel1 = get_data(d1)
            x2, y2, xlabel2 = get_data(d2)

            all_x = sorted(set(x1) | set(x2))
            y1_dict = dict(zip(x1, y1))
            y2_dict = dict(zip(x2, y2))
            y1_filled = [y1_dict.get(k, 0) for k in all_x]
            y2_filled = [y2_dict.get(k, 0) for k in all_x]
            plt.figure()
            if chart_type == "条形统计图":
                width = 0.35
                indices = list(range(len(all_x)))
                # 新增：用变量接收两组bar对象，分别标注
                bars1 = plt.bar([i - width / 2 for i in indices], y1_filled, width=width, label=d1)
                bars2 = plt.bar([i + width / 2 for i in indices], y2_filled, width=width, label=d2)
                # 新增：为第一组柱子（d1）标注金额
                for bar in bars1:
                    height = bar.get_height()
                    plt.text(bar.get_x() + bar.get_width()/2., height + 0.5,
                             f'{height:.2f}', ha='center', va='bottom')
                # 新增：为第二组柱子（d2）标注金额
                for bar in bars2:
                    height = bar.get_height()
                    plt.text(bar.get_x() + bar.get_width()/2., height + 0.5,
                             f'{height:.2f}', ha='center', va='bottom')
                plt.xticks(indices, all_x)
            elif chart_type == "折线统计图":
                # 新增：用变量接收两组line对象，分别标注
                line1 = plt.plot(all_x, y1_filled, label=d1, marker='o')[0]
                line2 = plt.plot(all_x, y2_filled, label=d2, marker='s')[0]
                # 新增：为第一组折线点（d1）标注金额
                for xi, yi in zip(all_x, y1_filled):
                    plt.text(xi, yi + 0.5, f'{yi:.2f}', ha='center', va='bottom')
                # 新增：为第二组折线点（d2）标注金额
                for xi, yi in zip(all_x, y2_filled):
                    plt.text(xi, yi + 0.5, f'{yi:.2f}', ha='center', va='bottom')
            else:
                return messagebox.showinfo("提示", "复合图暂不支持饼图")
            plt.legend()
            plt.title(f"{d1} vs {d2} 实付对比")
            plt.xlabel(xlabel1)
            plt.ylabel("实付总和")
            plt.tight_layout()
            plt.show()
        return

class InventoryPage:
    def __init__(self, parent):
        self.frame = tk.Frame(parent)
        configure_scaling_and_font(self.frame)

        # 顶部操作区
        top_frame = tk.Frame(self.frame)
        top_frame.pack(fill=tk.X, pady=10, padx=5)

        # 搜索框
        ttk.Label(top_frame, text="搜索商品:").pack(side=tk.LEFT, padx=5)
        self.search_var = tk.StringVar()
        search_entry = ttk.Entry(top_frame, textvariable=self.search_var, width=30)
        search_entry.pack(side=tk.LEFT, padx=5)
        search_btn = ttk.Button(top_frame, text="搜索", command=self.search_inventory)
        search_btn.pack(side=tk.LEFT, padx=5)
        show_all_btn = ttk.Button(top_frame, text="显示全部", command=self.refresh_inventory)
        show_all_btn.pack(side=tk.LEFT, padx=5)

        # 操作按钮
        add_btn = ttk.Button(top_frame, text="新增商品", command=self.add_item)
        add_btn.pack(side=tk.LEFT, padx=5)
        edit_btn = ttk.Button(top_frame, text="修改商品", command=self.edit_item)
        edit_btn.pack(side=tk.LEFT, padx=5)
        delete_btn = ttk.Button(top_frame, text="删除商品", command=self.delete_item)
        delete_btn.pack(side=tk.LEFT, padx=5)
        export_btn = ttk.Button(top_frame, text="导出库存表", command=self.export_inventory)
        export_btn.pack(side=tk.LEFT, padx=5)

        # 库存表格
        cols = ("类别", "名称", "售价", "会员价", "备注")
        self.tree, container = make_table(self.frame, cols)
        container.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)

        # 库存统计
        self.stats_frame = tk.Frame(self.frame)
        self.stats_frame.pack(fill=tk.X, pady=5, padx=5)
        self.total_items_label = ttk.Label(self.stats_frame, text="商品总数: 0")
        self.total_items_label.pack(side=tk.LEFT, padx=20)

        # 初始化加载数据
        self.refresh_inventory()

    def refresh_inventory(self):
        """刷新库存表格数据"""
        # 清空表格
        for item in self.tree.get_children():
            self.tree.delete(item)

        # 从数据库加载数据
        try:
            self.tree.delete(*self.tree.get_children())
            conn2 = sqlite3.connect(DB_PATH)
            c2 = conn2.cursor()
            # 查询语句中移除stock字段
            c2.execute("SELECT category, name, price, member_price, remark FROM inventory ORDER BY category, name")
            rows = c2.fetchall()
            conn2.close()

            for row in rows:
                # 插入数据时不再包含库存信息
                self.tree.insert('', tk.END, values=(
                    row[0],  # 类别
                    row[1],  # 名称
                    f"{row[2]:.2f}",  # 售价
                    f"{row[3]:.2f}",  # 会员价
                    row[4] or ""  # 备注
                ))

            # 更新统计信息
            self.update_inventory_stats()

        except Exception as e:
            messagebox.showerror("错误", f"加载库存数据失败: {str(e)}")

    def search_inventory(self):
        """搜索库存商品"""
        keyword = self.search_var.get().strip().lower()
        if not keyword:
            self.refresh_inventory()
            return

        # 清空表格
        for item in self.tree.get_children():
            self.tree.delete(item)

        try:
            conn = sqlite3.connect(DB_PATH)
            cursor = conn.cursor()
            cursor.execute("""
                SELECT category, name, price, member_price, remark 
                FROM inventory 
                WHERE category LIKE ? OR name LIKE ? OR remark LIKE ?
                ORDER BY category, name
            """, (f"%{keyword}%", f"%{keyword}%", f"%{keyword}%"))
            items = cursor.fetchall()
            conn.close()

            # 填充表格
            for item in items:
                formatted_item = (
                    item[0],  # 类别
                    item[1],  # 名称
                    f"{item[2]:.2f}",  # 售价
                    f"{item[3]:.2f}",  # 会员价
                    item[4] or ""  # 备注
                )
                self.tree.insert('', tk.END, values=formatted_item)

        except Exception as e:
            messagebox.showerror("错误", f"搜索库存失败: {str(e)}")

    def update_inventory_stats(self):
        """更新库存统计信息"""
        try:
            conn = sqlite3.connect(DB_PATH)
            cursor = conn.cursor()

            # 总商品数
            cursor.execute("SELECT COUNT(*) FROM inventory")
            total = cursor.fetchone()[0]
            conn.close()
            self.total_items_label.config(text=f"商品总数: {total}")
        except Exception as e:
            print(f"更新库存统计失败: {str(e)}")

    def add_item(self):
        top = tk.Toplevel(self.frame)
        top.title("新增商品")
        top.transient(self.frame)
        configure_scaling_and_font(top)

        tk.Label(top, text="类别:").grid(row=0, column=0, padx=5, pady=5, sticky='e')
        cat_e = tk.Entry(top)
        cat_e.grid(row=0, column=1)

        tk.Label(top, text="名称:").grid(row=1, column=0, padx=5, pady=5, sticky='e')
        name_e = tk.Entry(top)
        name_e.grid(row=1, column=1)

        tk.Label(top, text="售价:").grid(row=2, column=0, padx=5, pady=5, sticky='e')
        price_e = tk.Entry(top)
        price_e.grid(row=2, column=1)

        tk.Label(top, text="会员价:").grid(row=3, column=0, padx=5, pady=5, sticky='e')
        mem_price_e = tk.Entry(top)
        mem_price_e.grid(row=3, column=1)

        # 移除库存相关输入项
        # 原库存输入代码:
        # tk.Label(top, text="库存:").grid(row=4, column=0, padx=5, pady=5, sticky='e')
        # stock_e = tk.Entry(top)
        # stock_e.grid(row=4, column=1)

        tk.Label(top, text="备注:").grid(row=4, column=0, padx=5, pady=5, sticky='e')  # 调整行号
        rem_e = tk.Entry(top)
        rem_e.grid(row=4, column=1)  # 调整行号

        def save():
            # 处理逻辑中移除库存相关代码
            cat = cat_e.get().strip()
            name = name_e.get().strip()
            try:
                price = float(price_e.get())
                mem_price = float(mem_price_e.get())
            except:
                return messagebox.showerror("错误", "价格格式错误")

            rem = rem_e.get().strip()

            if not cat or not name:
                return messagebox.showerror("错误", "类别和名称不能为空")

            conn2 = sqlite3.connect(DB_PATH)
            c2 = conn2.cursor()
            try:
                # 插入语句中移除stock字段
                c2.execute("""
                    INSERT INTO inventory (category, name, price, member_price, remark) 
                    VALUES (?, ?, ?, ?, ?)
                """, (cat, name, price, mem_price, rem))
                conn2.commit()
                messagebox.showinfo("成功", f"商品 {name} 添加成功")
            except sqlite3.IntegrityError:
                conn2.close()
                return messagebox.showerror("错误", "商品名称已存在")
            except Exception as e:
                conn2.close()
                return messagebox.showerror("错误", f"添加失败: {str(e)}")
            finally:
                conn2.close()
            self.refresh_inventory()
            top.destroy()

        ttk.Button(top, text="保存", command=save).grid(row=5, columnspan=2, pady=10)  # 调整行号
        # 绑定回车键到保存函数
        top.bind('<Return>', lambda event: save())

    def edit_item(self):
        sel = self.tree.selection()
        if not sel:
            return messagebox.showerror("错误", "请选择商品")

        # 获取选中项数据（注意：由于移除了库存列，索引需要调整）
        item = self.tree.item(sel[0])['values']
        old_name = item[1]  # 名称现在是第2列（索引1）

        top = tk.Toplevel(self.frame)
        top.title("修改商品")
        top.transient(self.frame)
        configure_scaling_and_font(top)

        # 界面元素与添加商品类似，移除库存相关输入
        tk.Label(top, text="类别:").grid(row=0, column=0, padx=5, pady=5, sticky='e')
        cat_e = tk.Entry(top)
        cat_e.grid(row=0, column=1)
        cat_e.insert(0, item[0])

        tk.Label(top, text="名称:").grid(row=1, column=0, padx=5, pady=5, sticky='e')
        name_e = tk.Entry(top)
        name_e.grid(row=1, column=1)
        name_e.insert(0, item[1])

        tk.Label(top, text="售价:").grid(row=2, column=0, padx=5, pady=5, sticky='e')
        price_e = tk.Entry(top)
        price_e.grid(row=2, column=1)
        price_e.insert(0, item[2])

        tk.Label(top, text="会员价:").grid(row=3, column=0, padx=5, pady=5, sticky='e')
        mem_price_e = tk.Entry(top)
        mem_price_e.grid(row=3, column=1)
        mem_price_e.insert(0, item[3])

        # 移除库存输入项
        # 原库存输入代码已删除

        tk.Label(top, text="备注:").grid(row=4, column=0, padx=5, pady=5, sticky='e')  # 调整行号
        rem_e = tk.Entry(top)
        rem_e.grid(row=4, column=1)  # 调整行号
        rem_e.insert(0, item[4] if len(item) > 4 else "")

        def save():
            # 保存逻辑中移除库存相关处理
            cat = cat_e.get().strip()
            name = name_e.get().strip()
            try:
                price = float(price_e.get())
                mem_price = float(mem_price_e.get())
            except:
                return messagebox.showerror("错误", "价格格式错误")

            rem = rem_e.get().strip()

            if not cat or not name:
                return messagebox.showerror("错误", "类别和名称不能为空")

            conn2 = sqlite3.connect(DB_PATH)
            c2 = conn2.cursor()
            try:
                # 1. 查询旧数据（用于对比差异）
                c2.execute("""
                    SELECT category, name, price, member_price, remark, rowid 
                    FROM inventory WHERE name=?
                """, (old_name,))
                old_data = c2.fetchone()
                if not old_data:
                    conn2.close()
                    return messagebox.showerror("错误", "未找到该商品记录")

                old_category, old_name_db, old_price, old_mem_price, old_remark, item_id = old_data

                # 2. 对比差异，记录修改内容
                changes = []
                if cat != old_category:
                    changes.append(f"类别从「{old_category}」改为「{cat}」")
                if name != old_name_db:
                    changes.append(f"名称从「{old_name_db}」改为「{name}」")
                if price != old_price:
                    changes.append(f"售价从「{old_price:.2f}」改为「{price:.2f}」")
                if mem_price != old_mem_price:
                    changes.append(f"会员价从「{old_mem_price:.2f}」改为「{mem_price:.2f}」")
                if rem != old_remark:
                    changes.append(f"备注从「{old_remark or '无'}」改为「{rem or '无'}」")

                # 3. 更新数据库
                c2.execute("""
                    UPDATE inventory SET category=?, name=?, price=?, member_price=?, remark=? 
                    WHERE name=?
                """, (cat, name, price, mem_price, rem, old_name))
                conn2.commit()

                # 4. 推送修改事件（包含具体差异）
                if changes:  # 只有有修改时才推送
                    from test_server import data_queue
                    update_data = {
                        "id": item_id,
                        "changes": changes,  # 具体修改内容列表
                        "new_data": {
                            "category": cat,
                            "name": name,
                            "price": price,
                            "member_price": mem_price,
                            "remark": rem
                        }
                    }
                    data_queue.put(("update_inventory", [update_data]))
                    print(f"📤 已推送库存修改事件: ID={item_id}, 变更={changes}")

                messagebox.showinfo("成功", f"商品 {name} 修改成功")
            except sqlite3.IntegrityError:
                conn2.close()
                return messagebox.showerror("错误", "商品名称已存在")
            except Exception as e:
                conn2.close()
                return messagebox.showerror("错误", f"修改失败: {str(e)}")
            finally:
                conn2.close()
            self.refresh_inventory()
            top.destroy()

        ttk.Button(top, text="保存", command=save).grid(row=5, columnspan=2, pady=10)  # 调整行号
        top.bind('<Return>', lambda event: save())

    def delete_item(self):
        """删除选中的商品"""
        selected = self.tree.selection()
        if not selected:
            messagebox.showerror("错误", "请先选择要删除的商品")
            return

        item_values = self.tree.item(selected[0])['values']
        item_name = item_values[1]

        # 确认删除
        confirm = messagebox.askyesno("确认", f"确定要删除商品 '{item_name}' 吗？")
        if not confirm:
            return

        try:
            conn = sqlite3.connect(DB_PATH)
            cursor = conn.cursor()
            # 先查询商品ID（假设表中有id字段，根据实际表结构调整）
            cursor.execute("SELECT rowid FROM inventory WHERE name=?", (item_name,))
            result = cursor.fetchone()
            if not result:
                conn.close()
                return messagebox.showerror("错误", "未找到该商品记录")
            item_id = result[0]
            cursor.execute("DELETE FROM inventory WHERE name=?", (item_name,))
            conn.commit()
            conn.close()
            # 推送删除事件到网页
            from test_server import data_queue
            # 推送库存删除事件
            data_queue.put(("delete_inventory", [{"id": item_id, "name": item_name}]))
            print(f"📤 已推送库存删除事件: {item_name} (ID: {item_id})")

            # 推送联合刷新事件确保网页同步
            from datetime import datetime
            combined_payload = {
                "sales": [],
                "inventory": [],
                "accounting": [],
                "push_time": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                "full_refresh": True
            }
            data_queue.put(("combined", combined_payload))
            print(f"📤 已推送联合刷新事件")
            messagebox.showinfo("成功", f"商品 '{item_name}' 已删除")
            self.refresh_inventory()
        except Exception as e:
            messagebox.showerror("错误", f"删除商品失败: {str(e)}")

    def export_inventory(self):
        """导出库存数据到Excel"""
        # 获取保存路径
        file_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel 文件", "*.xlsx"), ("所有文件", "*.*")],
            title="导出库存数据"
        )

        if not file_path:
            return  # 用户取消导出

        try:
            # 创建Excel工作簿
            wb = Workbook()
            ws = wb.active
            ws.title = "库存数据"

            # 添加表头
            ws.append(["类别", "名称", "售价", "会员价", "备注"])

            # 添加数据
            for item in self.tree.get_children():
                values = self.tree.item(item)['values']
                ws.append(values)

            # 保存文件
            wb.save(file_path)
            messagebox.showinfo("成功", f"库存数据已成功导出到:\n{file_path}")

        except Exception as e:
            messagebox.showerror("错误", f"导出库存失败: {str(e)}")

class SettingsPage:
    def __init__(self, master, main_app):
        self.db_conn = None
        self.frame = tk.Frame(master)
        configure_scaling_and_font(self.frame)
        self.main_app = main_app
        # -------------------- 分类框架区域 --------------------

        # 用户管理区域
        user_frame = ttk.LabelFrame(self.frame, text="用户管理")
        user_frame.pack(fill=tk.X, padx=10, pady=5)
        ttk.Button(user_frame, text="新增用户", command=self.add_user).pack(side=tk.LEFT, padx=5, pady=5)
        ttk.Button(user_frame, text="修改用户", command=self.change_password).pack(side=tk.LEFT, padx=5, pady=5)
        ttk.Button(user_frame, text="删除用户", command=self.delete_user).pack(side=tk.LEFT, padx=5, pady=5)

        # 数据管理区域
        data_frame = ttk.LabelFrame(self.frame, text="数据管理")
        data_frame.pack(fill=tk.X, padx=10, pady=5)
        ttk.Button(data_frame, text="清空所有表格数据", command=self.clear_all).pack(side=tk.LEFT, padx=5, pady=5)
        ttk.Button(data_frame, text="推送所有数据", command=self.push_all_data).pack(side=tk.LEFT, padx=5, pady=5)

        # 备份与恢复区域
        backup_frame = ttk.LabelFrame(self.frame, text="备份与恢复")
        backup_frame.pack(fill=tk.X, padx=10, pady=5)
        ttk.Button(backup_frame, text="手动备份", command=self.manual_backup).pack(side=tk.LEFT, padx=5, pady=5)
        ttk.Button(backup_frame, text="恢复备份", command=self.restore_backup).pack(side=tk.LEFT, padx=5, pady=5)
        ttk.Button(backup_frame, text="备份设置", command=self.set_backup_settings).pack(side=tk.LEFT, padx=5, pady=5)

        # 系统设置区域
        system_frame = ttk.LabelFrame(self.frame, text="系统设置")
        system_frame.pack(fill=tk.X, padx=10, pady=5)
        ttk.Button(system_frame, text="定时提醒内存", command=self.set_memory_reminder).pack(side=tk.LEFT, padx=5,
                                                                                             pady=5)
        ttk.Button(system_frame, text="重置退出选项", command=self.reset_exit_settings).pack(side=tk.LEFT, padx=5,
                                                                                             pady=5)
        ttk.Button(system_frame, text="设置销售目标", command=self.set_sales_goals).pack(side=tk.LEFT, padx=5, pady=5)

        # 日志与更新区域
        log_frame = ttk.LabelFrame(self.frame, text="日志与更新")
        log_frame.pack(fill=tk.X, padx=10, pady=5)
        ttk.Button(log_frame, text="查看日志", command=self.show_log_window).pack(side=tk.LEFT, padx=5, pady=5)
        ttk.Button(log_frame, text="更新日志", command=self.show_update_log).pack(side=tk.LEFT, padx=5, pady=5)

        # 初始化日志缓冲区
        self.log_buffer = []
        # 重定向 stdout、stderr 和 logging 输出
        self.redirect_all_outputs()

        # ----------------- 以下为实例方法（注意都以 self 为首参数） -----------------


    def _create_progress_window(self, parent=None, title="操作中"):
        parent = parent or self.master
        win = tk.Toplevel(parent)
        win.title(title)
        win.resizable(False, False)
        win.transient(parent)
        win.grab_set()

        frm = ttk.Frame(win, padding=8)
        frm.pack(fill='both', expand=True)

        txt = scrolledtext.ScrolledText(frm, width=60, height=12, state='disabled')
        txt.pack(fill='both', expand=True, pady=(0, 8))

        pb = ttk.Progressbar(frm, mode='determinate', value=0)
        pb.pack(fill='x', pady=(0, 8))

        btn_frame = ttk.Frame(frm)
        btn_frame.pack(fill='x')
        cancel_flag = {'cancelled': False}

        def append_log(s):
            def _append():
                txt.config(state='normal')
                txt.insert('end', s + "\n")
                txt.yview_moveto(1.0)
                txt.config(state='disabled')

            win.after(0, _append)

        def set_progress(val, maximum=None):
            def _set():
                if maximum is not None:
                    pb.config(maximum=maximum)
                pb['value'] = val

            win.after(0, _set)

        def do_cancel():
            cancel_flag['cancelled'] = True
            append_log("用户请求取消操作，正在中断...")

        cancel_btn = ttk.Button(btn_frame, text="取消", command=do_cancel)
        cancel_btn.pack(side='right')

        return win, append_log, set_progress, cancel_flag

    def _make_reenable_callback(self, button):
        def _reenable():
            try:
                button.config(state='normal')
            except Exception:
                pass

        return _reenable

    def reset_exit_settings(self):
        """重置退出偏好设置"""
        reset_exit_preference()
        messagebox.showinfo("提示", "退出选项已重置，下次关闭时将显示选择窗口")

    def clear_all(self):
        if not messagebox.askyesno("确认", "确定要清空所有表格数据吗？此操作不可恢复。"):
            return
        if not messagebox.askyesno("二次确认", "请再次确认：是否清空所有表格数据？"):
            return
        try:
            conn2 = sqlite3.connect(DB_PATH)
            c2 = conn2.cursor()
            c2.execute("DELETE FROM inventory")
            c2.execute("DELETE FROM sales")
            c2.execute("DELETE FROM sale_items")
            c2.execute("DELETE FROM members")
            conn2.commit()
            conn2.close()
        except Exception as e:
            return messagebox.showerror("错误", f"清空失败：{e}")

        try:
            self.main_app.inventory_page.refresh_inventory()
            self.main_app.sales_page.refresh_today()
            self.main_app.members_page.refresh()
        except Exception as e:
            messagebox.showwarning("刷新失败", f"页面刷新失败：{e}")

        messagebox.showinfo("完成", "所有数据已清空，表格已刷新。")

    def push_all_data(self):
        try:
            # 直接用 DB_PATH 打开数据库文件
            conn = sqlite3.connect(DB_PATH)
            conn.row_factory = sqlite3.Row
            cur = conn.cursor()

            # 1) 全量读取 sales
            cur.execute("SELECT sale_id, datetime, total_due, total_paid, is_member FROM sales ORDER BY datetime")
            sales = [dict(r) for r in cur.fetchall()]

            # 2) 全量读取 inventory
            cur.execute("SELECT item_id, category, name, price, member_price, remark FROM inventory ORDER BY item_id")
            inventory = [dict(r) for r in cur.fetchall()]

            # 3) 全量读取 accounting
            cur.execute("SELECT sale_id, datetime, total_paid, is_member "
                        "FROM sales WHERE total_paid<=total_due ORDER BY datetime")
            accounting = [dict(r) for r in cur.fetchall()]

            conn.close()

            # 构造推送负载
            payload = {
                "sales": sales,
                "inventory": inventory,
                "accounting": accounting,
                "push_time": datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            }

            # 从配置文件读取公网地址。
            push_url = f"{CONFIG['push']['endpoint']}"
            resp = requests.post(push_url,
                                 json=payload, timeout=10)
            resp.raise_for_status()
            messagebox.showinfo("推送结果", "✅ 全部数据已成功推送！")
        except Exception as e:
            messagebox.showerror("推送失败", f"❌ {e}")

    def add_user(self):
        top = tk.Toplevel(self.frame)
        top.title("新增用户")
        top.transient(self.frame)
        configure_scaling_and_font(top)

        # 用户名
        tk.Label(top, text="用户名:").grid(row=0, column=0, padx=5, pady=5)
        user_e = tk.Entry(top)
        user_e.grid(row=0, column=1, padx=5, pady=5)

        # 密码
        tk.Label(top, text="密码:").grid(row=1, column=0, padx=5, pady=5)
        pwd_e = tk.Entry(top, show="*")
        pwd_e.grid(row=1, column=1, padx=5, pady=5)

        # 确认密码
        tk.Label(top, text="确认密码:").grid(row=2, column=0, padx=5, pady=5)
        cpwd_e = tk.Entry(top, show="*")
        cpwd_e.grid(row=2, column=1, padx=5, pady=5)

        # 新增头像选择
        tk.Label(top, text="头像:").grid(row=3, column=0, padx=5, pady=5)
        avatar_path_var = tk.StringVar()
        avatar_entry = ttk.Entry(top, textvariable=avatar_path_var, state='readonly', width=30)
        avatar_entry.grid(row=3, column=1, padx=5, pady=5)

        # 头像预览
        avatar_preview = tk.Label(top)
        avatar_preview.grid(row=0, column=2, rowspan=3, columnspan=1, padx=10, pady=10)

        # 修改选择头像的函数，添加裁剪功能
        def choose_avatar():
            """选择头像文件并裁剪为正方形"""
            file_path = filedialog.askopenfilename(
                title="选择头像",
                filetypes=[("图片文件", "*.png;*.jpg;*.jpeg;*.gif")]
            )
            if file_path:
                try:
                    # 打开图片检查是否为正方形
                    img = Image.open(file_path)
                    width, height = img.size

                    # 如果不是正方形，打开裁剪窗口
                    if width != height:
                        # 使用当前的top窗口作为父窗口
                        cropper = AvatarCropper(top, file_path)
                        top.wait_window(cropper.top)  # 等待裁剪窗口关闭

                        if cropper.cropped_image:
                            # 保存裁剪后的图片
                            import tempfile
                            temp_file = tempfile.NamedTemporaryFile(suffix=".png", delete=False)
                            cropper.cropped_image.save(temp_file, "PNG")
                            temp_file.close()
                            file_path = temp_file.name
                            img = cropper.cropped_image

                    # 更新头像路径和预览
                    avatar_path_var.set(file_path)
                    img.thumbnail((150, 150))
                    photo = ImageTk.PhotoImage(img)
                    avatar_preview.config(image=photo)
                    avatar_preview.image = photo
                except Exception as e:
                    print(f"处理头像失败: {e}")
                    messagebox.showerror("错误", "处理头像时发生错误")

        ttk.Button(top, text="选择头像", command=choose_avatar).grid(row=3, column=2, padx=5, pady=5)
        tk.Label(top, text="(不选择则使用默认头像)").grid(row=4, column=1, sticky='w', padx=5)

        def save():
            u = user_e.get().strip()
            p = pwd_e.get()
            cp = cpwd_e.get()
            if not u or not p or not cp:
                return messagebox.showerror("错误", "不能为空")
            if p != cp:
                return messagebox.showerror("错误", "密码不一致")

            # 获取头像路径，默认为"profile photo.png"
            avatar_path = avatar_path_var.get() or "profile photo.png"

            try:
                conn2 = sqlite3.connect(DB_PATH)
                c2 = conn2.cursor()
                # 插入用户信息，包括头像路径
                c2.execute("INSERT INTO users (username, password, avatar) VALUES (?, ?, ?)", (u, p, avatar_path))
                conn2.commit()
                conn2.close()
                messagebox.showinfo("完成", "用户已添加")
                top.destroy()
            except sqlite3.IntegrityError:
                messagebox.showerror("错误", "用户名已存在")

        ttk.Button(top, text="保存", command=save).grid(row=6, columnspan=3, pady=10)
        top.bind('<Return>', lambda event: save())

    def change_password(self):
        top = tk.Toplevel(self.frame)
        top.title("修改用户")
        top.transient(self.frame)
        configure_scaling_and_font(top)

        # 用户名选择
        tk.Label(top, text="用户名:").grid(row=0, column=0, padx=5, pady=5)
        conn2 = sqlite3.connect(DB_PATH)
        c2 = conn2.cursor()
        c2.execute("SELECT username FROM users")
        users = [r[0] for r in c2.fetchall()]
        conn2.close()
        user_var = tk.StringVar()
        user_cb = ttk.Combobox(top, values=users, textvariable=user_var, state="readonly")
        user_cb.grid(row=0, column=1, padx=5, pady=5)

        # 旧密码
        tk.Label(top, text="旧密码:").grid(row=1, column=0, padx=5, pady=5)
        old_e = tk.Entry(top, show="*")
        old_e.grid(row=1, column=1, padx=5, pady=5)

        # 新密码
        tk.Label(top, text="新密码:").grid(row=2, column=0, padx=5, pady=5)
        new_e = tk.Entry(top, show="*")
        new_e.grid(row=2, column=1, padx=5, pady=5)

        # 确认新密码
        tk.Label(top, text="确认新密码:").grid(row=3, column=0, padx=5, pady=5)
        cnew_e = tk.Entry(top, show="*")
        cnew_e.grid(row=3, column=1, padx=5, pady=5)

        # 头像设置
        tk.Label(top, text="头像:").grid(row=4, column=0, padx=5, pady=5)
        avatar_path_var = tk.StringVar()
        avatar_entry = ttk.Entry(top, textvariable=avatar_path_var, state='readonly', width=30)
        avatar_entry.grid(row=4, column=1, padx=5, pady=5)

        # 头像预览
        avatar_preview = tk.Label(top)
        avatar_preview.grid(row=1, column=2, rowspan=3, columnspan=1, padx=10, pady=10)

        # 修改选择头像的函数，添加裁剪功能
        def choose_avatar():
            """选择头像文件并裁剪为正方形"""
            file_path = filedialog.askopenfilename(
                title="选择头像",
                filetypes=[("图片文件", "*.png;*.jpg;*.jpeg;*.gif")]
            )
            if file_path:
                try:
                    # 打开图片检查是否为正方形
                    img = Image.open(file_path)
                    width, height = img.size

                    # 如果不是正方形，打开裁剪窗口
                    if width != height:
                        # 使用当前的top窗口作为父窗口
                        cropper = AvatarCropper(top, file_path)
                        top.wait_window(cropper.top)  # 等待裁剪窗口关闭

                        if cropper.cropped_image:
                            # 保存裁剪后的图片
                            import tempfile
                            temp_file = tempfile.NamedTemporaryFile(suffix=".png", delete=False)
                            cropper.cropped_image.save(temp_file, "PNG")
                            temp_file.close()
                            file_path = temp_file.name
                            img = cropper.cropped_image

                    # 更新头像路径和预览
                    avatar_path_var.set(file_path)
                    img.thumbnail((150, 150))
                    photo = ImageTk.PhotoImage(img)
                    avatar_preview.config(image=photo)
                    avatar_preview.image = photo
                except Exception as e:
                    print(f"处理头像失败: {e}")
                    messagebox.showerror("错误", "处理头像时发生错误")

        ttk.Button(top, text="选择头像", command=choose_avatar).grid(row=4, column=2, padx=5, pady=5)
        tk.Label(top, text="(不选择则保持当前头像)").grid(row=5, column=1, sticky='w', padx=5)

        def update_avatar_display(username):
            """更新选中用户的头像显示"""
            if not username:
                return

            try:
                conn = sqlite3.connect(DB_PATH)
                cur = conn.cursor()
                cur.execute("SELECT avatar FROM users WHERE username=?", (username,))
                result = cur.fetchone()
                conn.close()

                if result and result[0]:
                    avatar_path = result[0]
                else:
                    avatar_path = "profile photo.png"  # 默认头像

                # 检查文件是否存在
                if not os.path.exists(avatar_path):
                    avatar_path = resource_path("profile photo.png")

                avatar_path_var.set(avatar_path)

                # 加载并显示头像
                img = Image.open(avatar_path)
                img.thumbnail((150, 150))
                photo = ImageTk.PhotoImage(img)
                avatar_preview.config(image=photo)
                avatar_preview.image = photo

            except Exception as e:
                print(f"加载头像失败: {e}")

        # 绑定用户选择事件，更新头像显示
        user_cb.bind("<<ComboboxSelected>>", lambda e: update_avatar_display(user_var.get()))

        def save():
            u = user_var.get()
            old = old_e.get()
            new = new_e.get()
            cnew = cnew_e.get()
            if not u or not old or not new or not cnew:
                return messagebox.showerror("错误", "不能为空")

            conn3 = sqlite3.connect(DB_PATH)
            c3 = conn3.cursor()
            c3.execute("SELECT password FROM users WHERE username=?", (u,))
            pw = c3.fetchone()
            if not pw or pw[0] != old:
                conn3.close()
                return messagebox.showerror("错误", "旧密码错误")

            if new != cnew:
                conn3.close()
                return messagebox.showerror("错误", "新密码不一致")

            # 获取头像路径，空则使用当前头像
            avatar_path = avatar_path_var.get() or None
            if avatar_path:
                # 更新密码和头像
                c3.execute("UPDATE users SET password=?, avatar=? WHERE username=?", (new, avatar_path, u))
            else:
                # 只更新密码
                c3.execute("UPDATE users SET password=? WHERE username=?", (new, u))

            conn3.commit()
            conn3.close()
            messagebox.showinfo("完成", "用户设置已更新")
            top.destroy()

        ttk.Button(top, text="保存", command=save).grid(row=7, columnspan=3, pady=10)
        top.bind('<Return>', lambda event: save())

    def delete_user(self):
        top = tk.Toplevel(self.frame)
        top.title("删除用户")
        top.transient(self.frame)
        configure_scaling_and_font(top)

        # 用户名选择
        tk.Label(top, text="用户名:").grid(row=0, column=0, padx=5, pady=5)
        conn2 = sqlite3.connect(DB_PATH)
        c2 = conn2.cursor()
        c2.execute("SELECT username FROM users")
        users = [r[0] for r in c2.fetchall()]
        conn2.close()
        user_var = tk.StringVar()
        user_cb = ttk.Combobox(top, values=users, textvariable=user_var, state="readonly")
        user_cb.grid(row=0, column=1, padx=5, pady=5)

        # 密码验证
        tk.Label(top, text="密码:").grid(row=1, column=0, padx=5, pady=5)
        pwd_e = tk.Entry(top, show="*")
        pwd_e.grid(row=1, column=1, padx=5, pady=5)

        # 头像显示区域
        tk.Label(top, text="用户头像:").grid(row=2, column=0, padx=5, pady=5)
        avatar_preview = tk.Label(top)
        avatar_preview.grid(row=2, column=1, columnspan=3, pady=10)

        def update_avatar_display(username):
            """更新选中用户的头像显示"""
            if not username:
                return

            try:
                conn = sqlite3.connect(DB_PATH)
                cur = conn.cursor()
                cur.execute("SELECT avatar FROM users WHERE username=?", (username,))
                result = cur.fetchone()
                conn.close()

                if result and result[0]:
                    avatar_path = result[0]
                else:
                    avatar_path = "profile photo.png"  # 默认头像

                # 检查文件是否存在
                if not os.path.exists(avatar_path):
                    avatar_path = resource_path("profile photo.png")

                # 加载并显示头像
                img = Image.open(avatar_path)
                img.thumbnail((150, 150))
                photo = ImageTk.PhotoImage(img)
                avatar_preview.config(image=photo)
                avatar_preview.image = photo

            except Exception as e:
                print(f"加载头像失败: {e}")

        # 绑定用户选择事件，更新头像显示
        user_cb.bind("<<ComboboxSelected>>", lambda e: update_avatar_display(user_var.get()))

        def delete():
            u = user_var.get()
            p = pwd_e.get()
            conn3 = sqlite3.connect(DB_PATH)
            c3 = conn3.cursor()
            c3.execute("SELECT password FROM users WHERE username=?", (u,))
            pw = c3.fetchone()
            if not pw or pw[0] != p:
                conn3.close()
                return messagebox.showerror("错误", "密码错误")

            if messagebox.askyesno("确认", f"删除用户 {u}？"):
                c3.execute("DELETE FROM users WHERE username=?", (u,))
                conn3.commit()
                conn3.close()
                messagebox.showinfo("完成", "用户已删除")
                top.destroy()
            else:
                conn3.close()

        ttk.Button(top, text="删除", command=delete).grid(row=4, columnspan=2, pady=10)
        top.bind('<Return>', lambda event: delete())

    def set_memory_reminder(self):
        top = tk.Toplevel(self.frame)
        top.title("定时提醒内存")
        top.transient(self.frame)
        configure_scaling_and_font(top)
        tk.Label(top, text="每几个月提醒一次:").pack(pady=5)
        month_e = tk.Entry(top)
        month_e.pack(pady=5)

        def confirm():
            m = month_e.get().strip()
            if not m.isdigit():
                return messagebox.showerror("错误", "请输入数字")
            m = int(m)
            now = datetime.now().strftime("%Y-%m-%d")
            conn3 = sqlite3.connect(DB_PATH)
            c3 = conn3.cursor()
            c3.execute("SELECT COUNT(*) FROM memory_reminder")
            if c3.fetchone()[0] > 0:
                c3.execute("UPDATE memory_reminder SET last_reminder_date=?, reminder_interval=?",
                           (now, m))
            else:
                c3.execute("INSERT INTO memory_reminder (last_reminder_date, reminder_interval) VALUES (?,?)",
                           (now, m))
            conn3.commit()
            conn3.close()
            messagebox.showinfo("完成", f"每 {m} 月提醒一次已设置")
            top.destroy()

        ttk.Button(top, text="确认", command=confirm).pack(pady=10)
        top.bind('<Return>', lambda event: confirm())

    def redirect_all_outputs(self):
        """同时捕获 stdout、stderr 和 logging 日志"""
        # 保存原始输出
        self.original_stdout = sys.stdout
        self.original_stderr = sys.stderr

        # 1. 重定向 stdout 和 stderr 到自定义 Logger
        # 在 SettingsPage 类的 redirect_all_outputs 方法中修改 MultiLogger 类
        class MultiLogger:
            def __init__(self, outer):
                self.outer = outer  # 持有 SettingsPage 实例引用

            def write(self, text):
                if text.strip():  # 过滤空行
                    timestamp = datetime.now().strftime("%Y.%m.%d %H:%M:%S")
                    self.outer.log_buffer.append(f"{timestamp} ---- ➡️ {text.strip()}")
                    self.outer.save_logs_to_file()  # 保存到日志文件

                # 关键修复：仅在原始输出流存在时调用 write
                if self.outer.original_stdout is not None:
                    self.outer.original_stdout.write(text)
                # 可选：如果需要兼容 stderr，补充对 original_stderr 的判断
                # if self.outer.original_stderr is not None and self is sys.stderr:
                #     self.outer.original_stderr.write(text)

            def flush(self):
                # 关键修复：仅在原始输出流存在时调用 flush
                if self.outer.original_stdout is not None:
                    self.outer.original_stdout.flush()
                # 可选：兼容 stderr
                # if self.outer.original_stderr is not None and self is sys.stderr:
                #     self.outer.original_stderr.flush()

        # 替换 stdout 和 stderr
        sys.stdout = MultiLogger(self)
        sys.stderr = MultiLogger(self)

        # 2. 捕获 logging 模块输出（如Flask服务器日志）
        class LoggingHandler(logging.Handler):
            def __init__(self, outer):
                super().__init__()
                self.outer = outer

            def emit(self, record):
                # 格式化日志记录
                msg = self.format(record)
                timestamp = datetime.now().strftime("%Y.%m.%d %H:%M:%S")
                self.outer.log_buffer.append(f"{timestamp} ---- ➡️ {msg}")
                # 新增：将日志保存到文件
                self.outer.save_logs_to_file()

        # 配置 logging 捕获
        self.logging_handler = LoggingHandler(self)
        # 设置日志级别（捕获所有级别）
        self.logging_handler.setLevel(logging.DEBUG)
        # 添加到全局日志器
        logging.getLogger().addHandler(self.logging_handler)
        # 确保Flask的日志也被捕获
        logging.getLogger('werkzeug').addHandler(self.logging_handler)  # Flask服务器日志

    def init_log_directory(self):
        """初始化日志目录（根目录/SistersOS_log）"""
        # 获取程序根目录
        root_dir = os.path.dirname(sys.argv[0])
        log_dir = os.path.join(root_dir, "SistersOS_log")
        if not os.path.exists(log_dir):
            os.makedirs(log_dir)
        return log_dir

    def save_logs_to_file(self):
        """将日志保存到当天的文件中"""
        if not self.log_buffer:  # 如果日志缓冲区为空则不操作
            return

        try:
            log_dir = self.init_log_directory()
            # 生成当天日志文件名（年_月_日.txt）
            today = datetime.now().strftime("%Y_%m_%d")
            log_filename = f"{today}.txt"
            log_path = os.path.join(log_dir, log_filename)

            # 检查文件是否存在，不存在则写入头部信息
            file_exists = os.path.exists(log_path)

            # 只写入新增的日志（最后一条）
            with open(log_path, "a", encoding="utf-8") as f:
                if not file_exists:
                    f.write("系统日志记录：\n")
                    f.write("------------------------\n")
                # 写入最后一条日志（因为日志是实时捕获的，每次只需要写入最新的）
                f.write(self.log_buffer[-1] + "\n")

            # 检查并清理旧日志
            self.cleanup_old_logs(log_dir)
        except Exception as e:
            print(f"日志保存失败: {str(e)}")

    def cleanup_old_logs(self, log_dir, max_keep=30):
        """清理旧日志，只保留最近30个文件"""
        try:
            # 获取所有日志文件并按创建时间排序（最早的在前面）
            log_files = glob.glob(os.path.join(log_dir, "*.txt"))
            # 按创建时间排序， oldest first
            log_files.sort(key=lambda x: os.path.getctime(x))

            # 如果超过最大保留数量，删除最早的
            if len(log_files) > max_keep:
                files_to_delete = log_files[:len(log_files) - max_keep]
                for file in files_to_delete:
                    os.remove(file)
        except Exception as e:
            print(f"日志清理失败: {str(e)}")

    def show_log_window(self):
        """显示日志窗口（保持原有逻辑，无需修改）"""
        log_win = tk.Toplevel(self.frame)
        log_win.title("系统日志")
        log_win.geometry(f"800x500+{self.frame.winfo_x() + 50}+{self.frame.winfo_y() + 50}")
        configure_scaling_and_font(log_win)
        # 导出按钮
        export_btn = ttk.Button(log_win, text="导出日志", command=self.export_logs)
        export_btn.pack(pady=10)

        # 日志显示区域
        log_frame = ttk.Frame(log_win)
        log_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

        self.log_text = scrolledtext.ScrolledText(
            log_frame,
            wrap=tk.WORD,
            font=("Microsoft YaHei", int(9 * SCALE_FACTOR))
        )
        self.log_text.pack(fill=tk.BOTH, expand=True)

        # 加载日志内容
        self.log_text.insert(tk.END, "系统日志记录：\n")
        self.log_text.insert(tk.END, "------------------------\n")
        for log in self.log_buffer:
            self.log_text.insert(tk.END, log + "\n")
        self.log_text.see(tk.END)  # 滚动到底部

        # 实时更新日志
        def update_log():
            current_len = len(self.log_buffer)
            # 只添加新增的日志
            if current_len > len(self.log_text.get(1.0, tk.END).splitlines()) - 2:
                new_logs = self.log_buffer[-(current_len - (len(self.log_text.get(1.0, tk.END).splitlines()) - 2)):]
                for log in new_logs:
                    self.log_text.insert(tk.END, log + "\n")
                self.log_text.see(tk.END)
            log_win.after(1000, update_log)  # 每秒更新一次

        update_log()

    def export_logs(self):
        """导出日志为TXT（保持原有逻辑）"""
        file_path = filedialog.asksaveasfilename(
            defaultextension=".txt",
            filetypes=[("文本文件", "*.txt"), ("所有文件", "*.*")],
            title="导出日志"
        )
        if file_path:
            with open(file_path, 'w', encoding='utf-8') as f:
                f.write("系统日志记录：\n")
                f.write("------------------------\n")
                for log in self.log_buffer:
                    f.write(log + "\n")
            messagebox.showinfo("成功", f"日志已导出至：\n{file_path}")

    def manual_backup(self):
        """全局手动备份函数（供调度线程调用）"""
        try:
            backup_path = CONFIG["backup"]["path"]
            if not backup_path:
                backup_path = os.path.join(os.path.dirname(DB_PATH), "backups")
            os.makedirs(backup_path, exist_ok=True)

            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            backup_file = os.path.join(backup_path, f"backup_{timestamp}.db")
            shutil.copy2(DB_PATH, backup_file)

            print(f"[{datetime.now().strftime('%Y.%m.%d %H:%M:%S')}] 自动备份成功：{backup_file}")
            messagebox.showinfo("备份成功", f"备份已完成：\n{backup_file}")
            return True
        except Exception as e:
            error_msg = f"[{datetime.now().strftime('%Y.%m.%d %H:%M:%S')}] 自动备份失败：{str(e)}"
            print(error_msg)
            messagebox.showerror("备份失败", f"备份过程出错：\n{str(e)}")
            return False

    def restore_backup(self):
        """从备份恢复数据库"""
        try:
            # 让用户选择备份文件
            file_path = filedialog.askopenfilename(
                filetypes=[("数据库文件", "*.db"), ("所有文件", "*.*")],
                initialdir=CONFIG["backup"]["path"],
                title="选择备份文件"
            )

            if not file_path:
                return False

            # 确认恢复操作
            if not messagebox.askyesno("确认恢复",
                                       f"确定要从以下备份恢复吗？\n{file_path}\n\n此操作将覆盖当前数据！"):
                return False

            # 先关闭所有数据库连接并备份当前数据
            temp_backup = f"{DB_PATH}.temp_backup"
            shutil.copy2(DB_PATH, temp_backup)

            try:
                # 执行恢复
                shutil.copy2(file_path, DB_PATH)

                # 刷新所有页面数据
                self.main_app.inventory_page.refresh_inventory()
                self.main_app.sales_page.refresh_today()
                self.main_app.members_page.refresh()

                messagebox.showinfo("成功", "数据已从备份恢复")
                os.remove(temp_backup)  # 恢复成功则删除临时备份
                return True
            except Exception as e:
                # 恢复失败则回滚到临时备份
                shutil.copy2(temp_backup, DB_PATH)
                os.remove(temp_backup)
                raise e

        except Exception as e:
            messagebox.showerror("失败", f"恢复失败:\n{str(e)}")
        return False

    def set_backup_settings(self):
        """备份设置窗口（新增分钟间隔配置）"""
        top = tk.Toplevel(self.frame)
        top.title("备份设置")
        top.transient(self.frame)
        configure_scaling_and_font(top)

        # 启用自动备份
        tk.Label(top, text="启用自动备份:").grid(row=0, column=0, padx=10, pady=10, sticky='e')
        enable_var = tk.BooleanVar(value=CONFIG["backup"]["enabled"].lower() == "true")
        ttk.Checkbutton(top, variable=enable_var).grid(row=0, column=1, pady=10, sticky='w')

        # 备份分钟间隔（新增）
        tk.Label(top, text="备份间隔(分钟):").grid(row=1, column=0, padx=10, pady=10, sticky='e')
        interval_var = tk.StringVar(value=CONFIG["backup"]["interval_minutes"])
        ttk.Entry(top, textvariable=interval_var, width=10).grid(row=1, column=1, pady=10)
        tk.Label(top, text="(每次备份后间隔N分钟)").grid(row=1, column=2, padx=5, sticky='w')

        # 备份路径（保持不变）
        tk.Label(top, text="备份保存路径:").grid(row=2, column=0, padx=10, pady=10, sticky='e')
        path_var = tk.StringVar(value=CONFIG["backup"]["path"])
        ttk.Entry(top, textvariable=path_var, width=20).grid(row=2, column=1, pady=10)

        def choose_path():
            path = filedialog.askdirectory(title="选择备份路径")
            if path:
                path_var.set(path)

        ttk.Button(top, text="浏览...", command=choose_path).grid(row=2, column=2, padx=5)

        # 保存设置（更新配置项）
        def save_settings():
            try:
                # 验证分钟间隔为正整数
                interval = int(interval_var.get().strip())
                if interval <= 0:
                    raise ValueError("间隔必须为正整数")

                # 更新配置
                CONFIG["backup"]["enabled"] = "true" if enable_var.get() else "false"
                CONFIG["backup"]["interval_minutes"] = str(interval)  # 保存分钟间隔
                CONFIG["backup"]["path"] = os.path.abspath(path_var.get())  # 转为绝对路径

                save_config_to_file()  # 保存到文件
                self.log_buffer.append(
                    f"[{datetime.now().strftime('%Y.%m.%d %H:%M:%S')}] 备份设置已更新（间隔{interval}分钟）")
                messagebox.showinfo("成功", "备份设置已保存")
                top.destroy()
            except ValueError as e:
                messagebox.showerror("输入错误", f"分钟间隔无效：{str(e)}")
            except Exception as e:
                messagebox.showerror("失败", f"保存设置失败：{str(e)}")

        ttk.Button(top, text="保存设置", command=save_settings).grid(row=3, column=0, columnspan=3, pady=20)

    def choose_backup_path(self, path_var):
        """选择备份保存路径"""
        path = filedialog.askdirectory(title="选择备份保存路径")
        if path:
            path_var.set(path)

    def show_update_log(self):
        """读取并显示历代版本更新说明"""
        try:
            # 获取日志文件路径（根目录下的txt文件）
            log_file_path = resource_path("历代版本更新说明.txt")  # 使用已有的resource_path工具函数

            # 检查文件是否存在
            if not os.path.exists(log_file_path):
                messagebox.showerror("错误", "未找到更新日志文件\n请确保根目录下存在「历代版本更新说明.txt」")
                return

            # 读取文件内容（支持UTF-8编码）
            with open(log_file_path, 'r', encoding='utf-8') as f:
                log_content = f.read()

            # 创建显示窗口
            log_window = tk.Toplevel(self.frame)
            log_window.title("历代版本更新说明")
            log_window.geometry("1000x600")
            log_window.transient(self.frame)  # 依附于主窗口
            log_window.grab_set()  # 模态窗口

            # 添加滚动条
            scrollbar = ttk.Scrollbar(log_window)
            scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

            # 文本显示区域（只读）
            text_widget = tk.Text(
                log_window,
                wrap=tk.WORD,
                yscrollcommand=scrollbar.set,
                font=('Microsoft YaHei', int(10 * SCALE_FACTOR)),
                padx=10,
                pady=10
            )
            text_widget.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
            text_widget.insert(tk.END, log_content)
            text_widget.config(state=tk.DISABLED)  # 设置为只读

            # 关联滚动条
            scrollbar.config(command=text_widget.yview)

            # 关闭按钮
            ttk.Button(
                log_window,
                text="关闭",
                command=log_window.destroy
            ).pack(pady=10)

        except Exception as e:
            messagebox.showerror("读取失败", f"日志文件读取错误：\n{str(e)}")

    def set_sales_goals(self):
        """设置销售目标的窗口"""
        top = tk.Toplevel(self.frame)
        top.title("设置销售目标")
        top.transient(self.frame)
        configure_scaling_and_font(top)

        today = date.today()

        # 今日目标设置
        tk.Label(top, text="每日销售目标 (¥):").grid(row=0, column=0, padx=10, pady=15, sticky='e')
        day_date_str = today.strftime("%Y-%m-%d")
        day_goal = get_sales_goal("day", day_date_str)
        self.day_goal_var = tk.StringVar(value=str(day_goal))
        ttk.Entry(top, textvariable=self.day_goal_var, width=15).grid(row=0, column=1, pady=15)

        # 本月目标设置
        tk.Label(top, text="本月销售目标 (¥):").grid(row=1, column=0, padx=10, pady=15, sticky='e')
        month_date_str = today.strftime("%Y-%m")
        month_goal = get_sales_goal("month", month_date_str)
        self.month_goal_var = tk.StringVar(value=str(month_goal))
        ttk.Entry(top, textvariable=self.month_goal_var, width=15).grid(row=1, column=1, pady=15)

        # 目标历史记录按钮
        ttk.Button(top, text="查看历史目标", command=self.view_goal_history).grid(
            row=2, column=0, columnspan=2, pady=10)

        # 保存按钮
        def save_goals():
            try:
                # 验证输入
                day_amount = float(self.day_goal_var.get())
                month_amount = float(self.month_goal_var.get())

                if day_amount <= 0 or month_amount <= 0:
                    return messagebox.showerror("错误", "目标金额必须大于0")

                # 保存设置
                update_sales_goal("day", day_date_str, day_amount)
                update_sales_goal("month", month_date_str, month_amount)

                # 更新主界面进度显示
                self.main_app.update_sales_progress()

                messagebox.showinfo("成功", "销售目标已更新")
                top.destroy()
            except ValueError:
                messagebox.showerror("错误", "请输入有效的数字")

        ttk.Button(top, text="保存设置", command=save_goals).grid(
            row=3, column=0, columnspan=2, pady=20)
        top.bind('<Return>', lambda event: save_goals())

    def view_goal_history(self):
        """查看销售目标历史记录"""
        top = tk.Toplevel(self.frame)
        top.title("销售目标历史记录")
        top.transient(self.frame)
        configure_scaling_and_font(top)
        top.geometry("800x500")

        # 创建选项卡
        notebook = ttk.Notebook(top)
        notebook.pack(fill='both', expand=True, padx=10, pady=10)

        # 每日目标标签页
        day_frame = ttk.Frame(notebook)
        notebook.add(day_frame, text="每日目标")

        # 每月目标标签页
        month_frame = ttk.Frame(notebook)
        notebook.add(month_frame, text="每月目标")

        # 加载每日目标数据
        day_cols = ("日期", "目标金额(¥)", "实际销售额(¥)", "完成率")
        day_tree, day_container = make_table(day_frame, day_cols)
        day_container.pack(fill='both', expand=True)

        # 加载每月目标数据
        month_cols = ("月份", "目标金额(¥)", "实际销售额(¥)", "完成率")
        month_tree, month_container = make_table(month_frame, month_cols)
        month_container.pack(fill='both', expand=True)

        # 查询并填充数据
        conn = sqlite3.connect(DB_PATH)
        cur = conn.cursor()

        # 填充每日目标数据
        cur.execute("""
            SELECT target_date, target_amount FROM sales_goals 
            WHERE period_type='day' ORDER BY target_date DESC LIMIT 30
        """)
        day_goals = cur.fetchall()
        for date_str, goal in day_goals:
            # 获取该日的实际销售额
            cur.execute("""
                SELECT SUM(total_paid) FROM sales 
                WHERE datetime LIKE ?
            """, (f"{date_str}%",))
            sales = cur.fetchone()[0] or 0
            rate = (sales / goal) * 100 if goal > 0 else 0
            day_tree.insert('', 'end', values=(
                date_str,
                f"{goal:.2f}",
                f"{sales:.2f}",
                f"{rate:.1f}%"
            ))

        # 填充每月目标数据
        cur.execute("""
            SELECT target_date, target_amount FROM sales_goals 
            WHERE period_type='month' ORDER BY target_date DESC LIMIT 12
        """)
        month_goals = cur.fetchall()
        for date_str, goal in month_goals:
            # 获取该月的实际销售额
            cur.execute("""
                SELECT SUM(total_paid) FROM sales 
                WHERE datetime LIKE ?
            """, (f"{date_str}%",))
            sales = cur.fetchone()[0] or 0
            rate = (sales / goal) * 100 if goal > 0 else 0
            month_tree.insert('', 'end', values=(
                date_str,
                f"{goal:.2f}",
                f"{sales:.2f}",
                f"{rate:.1f}%"
            ))

        conn.close()

if __name__ == '__main__':
    # 单实例检查
    single_instance = SingleInstance()
    if single_instance.is_running():
        # 创建临时Tk实例显示提示（避免主线程冲突）
        temp_root = tk.Tk()
        temp_root.withdraw()  # 隐藏主窗口
        messagebox.showinfo("提示", "程序已在运行中！")
        temp_root.destroy()
        sys.exit(0)
    # 创建主窗口
    root = tk.Tk()

    # 1) 全局缩放与字体
    from tkinter import font as tkfont

    root.tk.call('tk', 'scaling', SCALE_FACTOR)
    base_font = tkfont.nametofont("TkDefaultFont")
    base_font.configure(family="Microsoft YaHei", size=int(10 * SCALE_FACTOR))
    style = ttk.Style()
    style.configure('.', font=(base_font.actual('family'), base_font.actual('size')))
    style.configure('Treeview', rowheight=int(24 * SCALE_FACTOR))

    # 2) 读取并应用主题
    selected = load_theme()
    if selected:
        try:
            style.theme_use(selected)
        except:
            pass

    # 3 确保警告被捕获
    warnings.filterwarnings("always")

    # 4 初始化数据库
    init_db()

    # 5 启动登录页
    app = LoginWindow(root)
    root.mainloop()